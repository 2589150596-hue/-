from openpyxl import load_workbook
path = r'C:\Users\Administrator\Desktop\汇总表_0609_1618.xlsx'
wb = load_workbook(path)
ws = wb['汇总表']

dispimg_count = 0
for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
    for cell in row:
        if cell.value and isinstance(cell.value, str) and 'DISPIMG' in cell.value:
            dispimg_count += 1
print('DISPIMG残留:', dispimg_count)

cell = ws['A1']
print('Header font:', cell.font.name, 'bold:', cell.font.bold)
print('Header fill:', cell.fill.start_color.rgb if cell.fill.start_color else None)

cell = ws['B2']
print('Border left:', cell.border.left.style)
print('Freeze:', ws.freeze_panes)
print('Row1 height:', ws.row_dimensions[1].height)
print('Row2 height:', ws.row_dimensions[2].height)
print('Col B width:', ws.column_dimensions['B'].width)
print('Col D width:', ws.column_dimensions['D'].width)

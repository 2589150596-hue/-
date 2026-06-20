import os
import re
import pandas as pd
from openpyxl import load_workbook
from collections import defaultdict

INPUT_DIR = r"C:\Users\Administrator\Desktop\2"
OUTPUT_FILE = r"C:\Users\Administrator\Desktop\2\统计结果.xlsx"
OUTPUT_FILE_ALT = r"D:\first-cc\统计结果.xlsx"  # 备用路径（当桌面文件被占用时）


def parse_customer(filename):
    """从文件名解析客户名: 序号-客户名-原名.xlsx"""
    name_without_ext = os.path.splitext(filename)[0]
    parts = name_without_ext.split("-", 2)
    if len(parts) >= 2:
        candidate = parts[1].strip()
        m = re.match(r"(\d+\.\d+)?(早补|晚补)?(.+)", candidate)
        if m:
            possible = m.group(3).strip()
            if possible:
                return possible
        return candidate
    return "未知客户"


def extract_shop_from_filename(filename):
    """从文件名提取店铺名（去掉序号、客户名、日期）。"""
    name_no_ext = os.path.splitext(filename)[0]
    parts = name_no_ext.split("-", 2)
    if len(parts) >= 3:
        raw = parts[2].strip()
        return clean_shop_name(raw)
    return "未知店铺"


def clean_shop_name(name):
    """清理店铺名称中的日期、序号、扩展名、括号等残留。"""
    if not name:
        return "未知店铺"
    name = str(name).strip()
    # 去掉 .xlsx 残留
    name = re.sub(r"\.xlsx", "", name, flags=re.IGNORECASE)
    # 去掉末尾日期模式: 6.10、(1)、(2)、-60-、-9- 等
    name = re.sub(r"\s*\(\d+\)\s*$", "", name)
    name = re.sub(r"-\d+-?\s*$", "", name)
    name = re.sub(r"\d+\.\d+.*$", "", name)
    name = re.sub(r"\d{4}[\.\s]+\d+[\.\s]+\d+.*$", "", name)
    # 去掉前后空格
    return name.strip()


def normalize_shop_name(name):
    """统一店铺名格式，用于合并重复项。"""
    if not name or str(name).strip() == "":
        return "未知店铺"
    name = str(name).strip()
    # 去掉全角和半角括号
    name = re.sub(r"[\(\)（）]", "", name)
    return name.strip()


def find_column_index(headers_r1, headers_r2, keywords):
    """按关键词列表查找列索引，支持两行表头。"""
    def search(headers):
        for keyword in keywords:
            for idx, h in enumerate(headers):
                if h and keyword in str(h).strip():
                    return idx
        return None

    idx = search(headers_r1)
    if idx is not None:
        return idx
    if headers_r2:
        idx = search(headers_r2)
        if idx is not None:
            return idx
    return None


def read_excel_data(filepath):
    """
    读取单个Excel，返回 (客户名, 总数量, 按店铺统计dict, 错误信息)。
    """
    filename = os.path.basename(filepath)
    customer = parse_customer(filename)
    shop_from_filename = extract_shop_from_filename(filename)

    try:
        wb = load_workbook(filepath, data_only=True)
        ws = wb.active
    except Exception as e:
        return customer, 0, {}, f"无法打开文件: {e}"

    # 读取前3行判断结构
    rows = list(ws.iter_rows(min_row=1, max_row=3, values_only=True))
    if not rows:
        return customer, 0, {}, "空文件"

    headers_r1 = [str(cell).strip() if cell is not None else None for cell in rows[0]]
    headers_r2 = None
    if len(rows) >= 2:
        has_header_keywords = any(
            cell is not None and isinstance(cell, str)
            and any(k in cell for k in ["数量", "价格", "时间"])
            for cell in rows[1]
        )
        if has_header_keywords:
            headers_r2 = [str(cell).strip() if cell is not None else None for cell in rows[1]]

    # 定位关键列
    qty_col = find_column_index(headers_r1, headers_r2, ["补单数量", "补数量", "下单数量", "数量"])
    if qty_col is None:
        return customer, 0, {}, f"未找到数量列"

    shop_col = find_column_index(headers_r1, headers_r2, ["店铺名称", "店铺", "客户(不填)", "客户名(不填)"])
    total_col = find_column_index(headers_r1, headers_r2, ["总计", "合计", "总"])

    data_start_row = 3 if headers_r2 else 2

    shop_stats = defaultdict(int)
    total = 0
    skipped_rows = 0

    for row in ws.iter_rows(min_row=data_start_row, values_only=True):
        if not row or all(cell is None for cell in row):
            continue

        # 跳过汇总行：总计=0 且 数量是数字
        if total_col is not None and total_col < len(row):
            total_val = row[total_col]
            qty_val_check = row[qty_col] if qty_col < len(row) else None
            if total_val == 0 and total_val is not False:
                try:
                    float(qty_val_check)
                    # 这是汇总行，跳过
                    continue
                except (ValueError, TypeError):
                    pass

        # 跳过含"合计"/"总计"/"汇总"字样的行
        for cell in row:
            if cell is not None and isinstance(cell, str) and any(k in cell for k in ["合计", "总计", "汇总"]):
                skipped_rows += 1
                break
        else:
            qty_val = row[qty_col] if qty_col < len(row) else None
            if qty_val is None:
                skipped_rows += 1
                continue
            try:
                qty = float(qty_val)
            except (ValueError, TypeError):
                skipped_rows += 1
                continue

            total += qty

            # 店铺名：表格里为空则用文件名提取的
            shop_name = None
            if shop_col is not None and shop_col < len(row):
                shop_name = row[shop_col]
            if not shop_name or str(shop_name).strip() == "":
                shop_name = shop_from_filename
            else:
                shop_name = clean_shop_name(shop_name)

            # 用规范化名称合并重复店铺
            normalized = normalize_shop_name(shop_name)
            shop_stats[normalized] += qty

    return customer, total, dict(shop_stats), None


def main():
    files = [f for f in os.listdir(INPUT_DIR)
             if f.lower().endswith(".xlsx") and not f.startswith("~$")
             and "统计结果" not in f and "统计" not in f]
    if not files:
        print(f"未在 {INPUT_DIR} 找到xlsx文件")
        return

    file_results = []
    customer_stats = defaultdict(float)
    shop_stats = defaultdict(float)

    for f in sorted(files, key=lambda x: int(re.findall(r"^\d+", x)[0]) if re.findall(r"^\d+", x) else 999):
        filepath = os.path.join(INPUT_DIR, f)
        customer, total, shops, err = read_excel_data(filepath)
        if err:
            file_results.append({"文件名": f, "客户名": customer, "补单数量": 0, "备注": err})
            print(f"[WARN] {f}: {err}")
        else:
            file_results.append({"文件名": f, "客户名": customer, "补单数量": total, "备注": ""})
            customer_stats[customer] += total
            for shop, qty in shops.items():
                shop_stats[shop] += qty
            print(f"[OK] {f}: 客户={customer}, 数量={total}")

    # 构建DataFrame
    df_file = pd.DataFrame(file_results)
    df_customer = pd.DataFrame(
        [{"客户名": k, "补单数量": v} for k, v in sorted(customer_stats.items(), key=lambda x: x[1], reverse=True)]
    )
    df_shop = pd.DataFrame(
        [{"店铺名": k, "补单数量": v} for k, v in sorted(shop_stats.items(), key=lambda x: x[1], reverse=True)]
    )

    # 导出Excel（先尝试桌面，被占用则写备用路径）
    out_path = OUTPUT_FILE
    try:
        with pd.ExcelWriter(OUTPUT_FILE, engine="openpyxl") as writer:
            df_file.to_excel(writer, sheet_name="按文件统计", index=False)
            df_customer.to_excel(writer, sheet_name="按客户统计", index=False)
            df_shop.to_excel(writer, sheet_name="按店铺统计", index=False)
    except PermissionError:
        out_path = OUTPUT_FILE_ALT
        with pd.ExcelWriter(OUTPUT_FILE_ALT, engine="openpyxl") as writer:
            df_file.to_excel(writer, sheet_name="按文件统计", index=False)
            df_customer.to_excel(writer, sheet_name="按客户统计", index=False)
            df_shop.to_excel(writer, sheet_name="按店铺统计", index=False)

    print(f"\n[统计完成] 结果已保存到: {out_path}")
    print(f"   文件数: {len(files)}, 客户数: {len(customer_stats)}, 店铺数: {len(shop_stats)}")


if __name__ == "__main__":
    main()

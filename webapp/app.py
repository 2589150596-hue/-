"""补单汇总工具 - Flask 网页应用"""
import os
import re
import uuid
import shutil
import json
import glob
import zipfile as zf
import io as io_mod
import time as time_mod
import threading
from datetime import datetime, timedelta, time, date
from flask import Flask, request, jsonify, send_file, render_template, Response
import numpy as np
import urllib.request as urllib_req

from utils import (
    read_excel_with_fallback,
    detect_abnormal_files,
    assign_shops_to_brushers,
    auto_schedule,
    interleave_shop_links,
    extract_customer_from_filename,
    extract_wps_images,
    extract_standard_images,
    extract_img_id,
    get_header_column_map,
    compute_stats,
    generate_output_excel,
)

app = Flask(__name__)
app.secret_key = os.urandom(24).hex()


@app.before_request
def handle_options():
    if request.method == 'OPTIONS':
        resp = app.make_default_options_response()
        resp.headers['Access-Control-Allow-Origin'] = '*'
        resp.headers['Access-Control-Allow-Methods'] = 'GET, POST, DELETE, OPTIONS'
        resp.headers['Access-Control-Allow-Headers'] = 'Content-Type'
        return resp


@app.after_request
def add_cors_headers(response):
    response.headers['Access-Control-Allow-Origin'] = '*'
    response.headers['Access-Control-Allow-Methods'] = 'GET, POST, DELETE, OPTIONS'
    response.headers['Access-Control-Allow-Headers'] = 'Content-Type'
    return response


def to_json_safe(df):
    """将DataFrame转换为JSON安全的dict列表"""
    records = df.fillna('').to_dict(orient='records')
    safe = []
    for row in records:
        safe_row = {}
        for k, v in row.items():
            if isinstance(v, (datetime, date, time)):
                safe_row[k] = str(v)
            elif isinstance(v, (np.integer,)):
                safe_row[k] = int(v)
            elif isinstance(v, (np.floating,)):
                safe_row[k] = float(v)
            elif isinstance(v, np.bool_):
                safe_row[k] = bool(v)
            elif v is None or (isinstance(v, float) and np.isnan(v)):
                safe_row[k] = ''
            else:
                safe_row[k] = v
        safe.append(safe_row)
    return safe

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
UPLOAD_DIR = os.path.join(BASE_DIR, 'uploads')
OUTPUT_DIR = os.path.join(BASE_DIR, 'outputs')
os.makedirs(UPLOAD_DIR, exist_ok=True)
os.makedirs(OUTPUT_DIR, exist_ok=True)

# 加载配置
CONFIG_FILE = os.path.join(BASE_DIR, 'config.json')
_config = {}
if os.path.isfile(CONFIG_FILE):
    with open(CONFIG_FILE, 'r', encoding='utf-8') as f:
        _config = json.load(f)


def _get_dingtalk_url():
    return _config.get('dingtalk_webhook', '')


def _send_dingtalk(msg):
    """发送钉钉机器人消息"""
    url = _get_dingtalk_url()
    if not url:
        return
    try:
        payload = {"msgtype": "text", "text": {"content": f"【补单汇总工具】{msg}"}}
        req = urllib_req.Request(url, data=json.dumps(payload).encode('utf-8'),
                                  headers={'Content-Type': 'application/json'})
        urllib_req.urlopen(req, timeout=5)
    except Exception:
        pass  # 钉钉发送失败不影响主流程


# SSE 实时刷新
_file_change_counter = 0
_sse_listeners = []


def _notify_change():
    global _file_change_counter
    _file_change_counter += 1


@app.route('/api/set-dingtalk', methods=['POST'])
def set_dingtalk():
    """设置钉钉机器人webhook地址"""
    data = request.get_json()
    url = data.get('url', '').strip()
    _config['dingtalk_webhook'] = url
    with open(CONFIG_FILE, 'w', encoding='utf-8') as f:
        json.dump(_config, f, ensure_ascii=False, indent=2)
    return jsonify({'ok': True, 'msg': '钉钉配置已保存'})


@app.route('/api/file-events')
def file_events():
    """SSE端点：文件变更时推送事件"""
    def stream():
        last = _file_change_counter
        while True:
            if _file_change_counter != last:
                last = _file_change_counter
                yield f"data: {last}\n\n"
            else:
                yield ":keepalive\n\n"
            time_mod.sleep(1)
    return Response(stream(), mimetype='text/event-stream',
                    headers={'Cache-Control': 'no-cache', 'Connection': 'keep-alive',
                             'X-Accel-Buffering': 'no'})

# ========== 路由 ==========

@app.route('/')
def index():
    return render_template('index.html')


def _next_serial(task_date_str):
    """获取指定任务日期的全局序号：统计所有文件中该日期的文件数+1（格式: MM.DD）"""
    prefix = task_date_str.replace('-', '.').split('.')  # "2026-06-23" → "6.23"
    if len(prefix) >= 3:
        date_prefix = f"{int(prefix[1])}.{int(prefix[2])}-"
    else:
        date_prefix = task_date_str + '-'
    count = 0
    for session_id in os.listdir(UPLOAD_DIR):
        sd = os.path.join(UPLOAD_DIR, session_id)
        if not os.path.isdir(sd):
            continue
        for ext in ('*.xlsx', '*.xls'):
            for f in glob.glob(os.path.join(sd, ext)):
                if os.path.basename(f).startswith(date_prefix):
                    count += 1
    return count + 1


def _load_processed(session_dir):
    """读取已处理文件列表"""
    pf = os.path.join(session_dir, 'processed.json')
    if os.path.isfile(pf):
        with open(pf, 'r', encoding='utf-8') as f:
            return set(json.load(f))
    return set()


def _mark_processed(session_dir, filenames):
    """标记文件为已处理"""
    processed = _load_processed(session_dir)
    processed.update(filenames)
    with open(os.path.join(session_dir, 'processed.json'), 'w', encoding='utf-8') as f:
        json.dump(list(processed), f)


_CN_NUMS = ['一', '二', '三', '四', '五', '六', '七', '八', '九', '十',
            '十一', '十二', '十三', '十四', '十五', '十六', '十七', '十八', '十九', '二十']


def _next_task_name():
    """生成下一个任务名：6月22日任务一.xlsx"""
    today = date.today()
    pattern = re.compile(rf'^{today.month}月{today.day}日任务([一二三四五六七八九十]+|\d+)\.xlsx$')
    max_n = 0
    for f in os.listdir(OUTPUT_DIR):
        m = pattern.match(f)
        if m:
            name = m.group(1)
            if name.isdigit():
                max_n = max(max_n, int(name))
            else:
                try:
                    max_n = max(max_n, _CN_NUMS.index(name) + 1)
                except ValueError:
                    pass
    n = max_n + 1
    cn = _CN_NUMS[n - 1] if n <= len(_CN_NUMS) else str(n)
    return f"{today.month}月{today.day}日任务{cn}.xlsx"


@app.route('/api/upload', methods=['POST'])
def upload():
    """顾客上传Excel文件，按 日期-序号-顾客名-源文件名 命名"""
    if 'files' not in request.files:
        return jsonify({'error': '未选择文件'}), 400

    files = request.files.getlist('files')
    customer_name = request.form.get('customer_name', '').strip()
    task_date_str = request.form.get('task_date', date.today().isoformat()).strip()
    if not customer_name:
        return jsonify({'error': '请填写顾客名称'}), 400

    # 格式化日期前缀：2026-06-23 → 6.23
    try:
        td = date.fromisoformat(task_date_str)
        date_prefix = f"{td.month}.{td.day}"
    except Exception:
        td = date.today()
        date_prefix = f"{td.month}.{td.day}"

    session_id = request.form.get('session_id', uuid.uuid4().hex)
    session_dir = os.path.join(UPLOAD_DIR, session_id)
    os.makedirs(session_dir, exist_ok=True)

    serial = _next_serial(task_date_str)
    uploaded = []
    for f in files:
        if f.filename and f.filename.lower().endswith(('.xlsx', '.xls')):
            orig_name = os.path.splitext(f.filename)[0]
            orig_clean = re.sub(r'^\d+[-_]', '', orig_name)
            new_name = f"{date_prefix}-{serial}-{customer_name}-{orig_clean}.xlsx"
            path = os.path.join(session_dir, new_name)
            f.save(path)
            uploaded.append(new_name)
            serial += 1

    _notify_change()
    total_msg = f"{customer_name} 上传了 {len(uploaded)} 个文件（任务日期：{date_prefix}）：{', '.join(uploaded)}"
    _send_dingtalk(total_msg)
    return jsonify({
        'session_id': session_id,
        'files': uploaded,
        'count': len(uploaded)
    })


@app.route('/api/files', methods=['GET'])
def list_files():
    """列出服务器上所有已上传文件"""
    all_files = []
    for session_id in os.listdir(UPLOAD_DIR):
        session_dir = os.path.join(UPLOAD_DIR, session_id)
        if not os.path.isdir(session_dir):
            continue
        for ext in ('*.xlsx', '*.xls'):
            for f in glob.glob(os.path.join(session_dir, ext)):
                basename = os.path.basename(f)
                if not basename.startswith('~$'):
                    all_files.append({
                        'session_id': session_id,
                        'filename': basename,
                        'path': f,
                        'mtime': os.path.getmtime(f)
                    })
    # 读取各session的已处理标记
    processed_map = {}
    for session_id in os.listdir(UPLOAD_DIR):
        sd = os.path.join(UPLOAD_DIR, session_id)
        if os.path.isdir(sd):
            processed_map[session_id] = _load_processed(sd)

    for f in all_files:
        f['processed'] = f['filename'] in processed_map.get(f['session_id'], set())
        # 从文件名解析任务日期（格式: MM.DD-序号-...）
        date_m = re.match(r'^(\d+)\.(\d+)-', f['filename'])
        if date_m:
            f['task_date'] = f"{date_m.group(1)}月{date_m.group(2)}日"
        else:
            f['task_date'] = '未知日期'

    all_files.sort(key=lambda x: (x.get('task_date', ''), x['session_id'], x['filename']))
    return jsonify({'files': all_files, 'count': len(all_files)})


@app.route('/api/download-file/<session_id>/<path:filename>')
def download_file(session_id, filename):
    """下载原始上传文件"""
    filepath = os.path.join(UPLOAD_DIR, session_id, filename)
    if not os.path.isfile(filepath):
        return jsonify({'error': '文件不存在'}), 404
    return send_file(filepath, as_attachment=True, download_name=filename)


@app.route('/api/delete-file', methods=['POST'])
def delete_file():
    """删除上传文件"""
    data = request.get_json()
    session_id = data.get('session_id', '')
    filename = data.get('filename', '')
    filepath = os.path.join(UPLOAD_DIR, session_id, filename)
    if not os.path.isfile(filepath):
        return jsonify({'error': '文件不存在'}), 404
    try:
        os.remove(filepath)
        _notify_change()
        return jsonify({'ok': True, 'filename': filename})
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/batch-download', methods=['POST'])
def batch_download():
    """批量下载选中的文件（打包为ZIP）"""
    data = request.get_json()
    files = data.get('files', [])
    if not files:
        return jsonify({'error': '未选择文件'}), 400

    buf = io_mod.BytesIO()
    with zf.ZipFile(buf, 'w', zf.ZIP_DEFLATED) as zipf:
        for item in files:
            filepath = os.path.join(UPLOAD_DIR, item['session_id'], item['filename'])
            if os.path.isfile(filepath):
                zipf.write(filepath, item['filename'])
    buf.seek(0)
    return send_file(buf, as_attachment=True, download_name='批量下载.zip',
                     mimetype='application/zip')


@app.route('/api/batch-delete', methods=['POST'])
def batch_delete():
    """批量删除选中的文件"""
    data = request.get_json()
    files = data.get('files', [])
    if not files:
        return jsonify({'error': '未选择文件'}), 400

    deleted, failed = [], []
    for item in files:
        filepath = os.path.join(UPLOAD_DIR, item['session_id'], item['filename'])
        if os.path.isfile(filepath):
            try:
                os.remove(filepath)
                deleted.append(item['filename'])
            except Exception as e:
                failed.append(item['filename'])
        else:
            failed.append(item['filename'])

    _notify_change()
    return jsonify({'ok': True, 'deleted': len(deleted), 'failed': failed})


@app.route('/api/process', methods=['POST'])
def process():
    """执行完整处理流程"""
    data = request.get_json()
    session_id = data.get('session_id', '')
    start_time = data.get('start_time', '08:00')
    end_time = data.get('end_time', '23:00')
    num_workers = int(data.get('num_workers', 1))
    min_interval = float(data.get('min_interval', 0))
    selected_files = data.get('selected_files', None)  # 员工自选的文件列表 [{session_id, filename}]

    # 收集所有可用文件
    if selected_files:
        # 员工手动选择了文件
        filepaths = []
        for item in selected_files:
            path = os.path.join(UPLOAD_DIR, item['session_id'], item['filename'])
            if os.path.isfile(path):
                filepaths.append(path)
    else:
        # 未选择时使用 session_id 目录下全部文件
        session_dir = os.path.join(UPLOAD_DIR, session_id)
        if not os.path.isdir(session_dir):
            return jsonify({'error': '会话已过期，请重新上传文件'}), 400
        filepaths = []
        for ext in ('*.xlsx', '*.xls'):
            filepaths.extend(glob.glob(os.path.join(session_dir, ext)))

    filepaths = [f for f in filepaths if not os.path.basename(f).startswith('~$')]

    if not filepaths:
        return jsonify({'error': '未找到Excel文件'}), 400

    import pandas as pd
    from openpyxl import load_workbook
    from openpyxl.drawing.image import Image as XLImage
    from openpyxl.utils import get_column_letter
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    import io

    try:
        # 1. 异常检测 — 有异常则终止处理
        abnormal = detect_abnormal_files(filepaths)
        has_issue = abnormal['duplicates'] or abnormal['invalid'] or abnormal.get('warnings')
        if has_issue:
            reasons = []
            for dup in abnormal['duplicates']:
                reasons.append(f"{dup[0]} 与 {dup[1]} 内容完全相同")
            for inv in abnormal['invalid']:
                reasons.append(f"{inv['file']}：{inv['reason']}")
            for warn in abnormal.get('warnings', []):
                reasons.append(f"{warn['file']}：{warn['reason']}")
            detail = '\n'.join(reasons)
            return jsonify({'error': f'发现 {len(reasons)} 个问题，已终止处理：\n{detail}\n\n请修正后重新上传。'}), 400

        # 2. 读取数据
        all_data = []
        customer_from_file = {}
        file_images = {}
        file_std_images = {}

        for f in filepaths:
            basename = os.path.basename(f)
            customer = extract_customer_from_filename(basename)
            customer_from_file[basename] = customer

            images = extract_wps_images(f)
            file_images[f] = images

            std_images_raw = extract_standard_images(f)
            std_images_mapped = {}
            if std_images_raw:
                header_map = get_header_column_map(f)
                for (row, col), img_data in std_images_raw.items():
                    std_col = header_map.get(col)
                    if std_col:
                        std_images_mapped.setdefault(std_col, {})[row] = img_data
                file_std_images[f] = std_images_mapped
            else:
                file_std_images[f] = {}

            df = read_excel_with_fallback(f)
            if len(df) == 0:
                continue

            if '客户(不填)' in df.columns:
                df['客户(不填)'] = df['客户(不填)'].astype(object)
                empty_mask = df['客户(不填)'].isna() | (df['客户(不填)'] == '')
                if empty_mask.any() and customer:
                    df.loc[empty_mask, '客户(不填)'] = customer
            elif customer:
                df['客户(不填)'] = customer

            df['来源文件'] = basename

            # 过滤无图片行
            if '__drawing_row__' in df.columns:
                pic_col = None
                for c in df.columns:
                    if str(c) == '主图':
                        pic_col = c
                        break
                if pic_col is not None:
                    valid_mask = []
                    for _, row in df.iterrows():
                        has_img = False
                        cell_val = row.get(pic_col)
                        if isinstance(cell_val, str):
                            img_id = extract_img_id(cell_val)
                            if img_id and img_id in images:
                                has_img = True
                        if not has_img:
                            dr = row.get('__drawing_row__')
                            if dr is not None:
                                dr_int = int(dr)
                                if pic_col in std_images_mapped and dr_int in std_images_mapped[pic_col]:
                                    has_img = True
                        valid_mask.append(has_img)
                    if not all(valid_mask):
                        df = df[valid_mask].reset_index(drop=True)

            all_data.append(df)

        combined = pd.concat(all_data, ignore_index=True)

        if '商品ID' in combined.columns:
            combined['商品ID'] = combined['商品ID'].apply(
                lambda x: str(int(x)) if pd.notna(x) and str(x).replace('.', '', 1).isdigit() else str(x) if pd.notna(x) else ''
            )

        if '备注' not in combined.columns:
            combined['备注'] = ''

        # 3. 识别关键列
        shop_col = None
        for col in combined.columns:
            if '店铺名称' in str(col) and combined[col].notna().sum() > 0:
                shop_col = col
                break
        if shop_col is None:
            for col in combined.columns:
                if ('客户' in str(col) and '店铺' in str(col)) or str(col) == '客户(不填)':
                    if combined[col].notna().sum() > 0:
                        shop_col = col
                        break
        if shop_col is None:
            for col in combined.columns:
                if str(col).startswith('Unnamed'):
                    continue
                if combined[col].apply(lambda x: isinstance(x, str) and len(x) > 3).sum() > 0:
                    shop_col = col
                    break

        qty_col = None
        for col in combined.columns:
            if '补单数量' in str(col) or '单量' in str(col) or '数量' in str(col):
                qty_col = col
                break

        if shop_col is None or qty_col is None:
            return jsonify({'error': '未识别到店铺列或单量列'}), 400

        # 4. 过滤 & 清理（保留数量为0的行，只过滤无店铺名或无数量值的空行）
        valid_mask = combined[shop_col].notna() & combined[qty_col].notna()
        data = combined[valid_mask].copy()
        data[qty_col] = pd.to_numeric(data[qty_col], errors='coerce').fillna(0)
        data[shop_col] = data[shop_col].astype(str).str.replace(r'[\n\r]+', '', regex=True).str.strip()

        # 5. 拆分 + 链接交叉
        expanded_rows = []
        for _, row in data.iterrows():
            qty = int(row[qty_col])
            if qty > 1:
                for _ in range(qty):
                    new_row = row.copy()
                    new_row[qty_col] = 1
                    expanded_rows.append(new_row)
            else:
                expanded_rows.append(row)
        data = pd.DataFrame(expanded_rows).reset_index(drop=True)
        data[qty_col] = pd.to_numeric(data[qty_col], errors='coerce')
        data = interleave_shop_links(data, shop_col)

        # 6. 调度
        shop_orders = data.groupby(shop_col)[qty_col].sum().astype(int).to_dict()
        if num_workers < 1:
            num_workers = 1

        schedule_list = auto_schedule(start_time, end_time, shop_orders, num_workers, min_interval)

        # 7. 映射回数据行
        shop_schedule_map = {}
        for item in schedule_list:
            shop_schedule_map.setdefault(item['店铺'], []).append(item)

        start_h, start_m = map(int, start_time.split(':'))
        base_time = datetime(2026, 1, 1, start_h, start_m)

        result_rows = []
        for shop_name, group in data.groupby(shop_col, sort=False):
            group = group.reset_index(drop=True)
            assignments_local = shop_schedule_map.get(shop_name, [])
            total_qty = len(group)
            for idx, (_, row) in enumerate(group.iterrows()):
                if idx < len(assignments_local):
                    item = assignments_local[idx]
                else:
                    item = {'时间分钟': 0, '刷手': '刷手1'}
                exec_time = base_time + timedelta(minutes=item['时间分钟'])
                new_row = row.to_dict()
                new_row['执行时间'] = exec_time.strftime('%H:%M')
                new_row['刷手'] = item['刷手']
                new_row['距离首单(分钟)'] = round(item['时间分钟'], 1)
                new_row['店铺总单量'] = total_qty
                result_rows.append(new_row)

        result_df = pd.DataFrame(result_rows)
        result_df = result_df.sort_values(by='执行时间').reset_index(drop=True)

        if '序号' in result_df.columns:
            result_df.drop(columns=['序号'], inplace=True)
        if '备注' not in result_df.columns:
            result_df['备注'] = ''

        # 列顺序
        cols = list(result_df.columns)
        ordered = ['执行时间', '刷手', '备注']
        unnamed_cols = [c for c in cols if str(c).startswith('Unnamed')]
        remaining = [c for c in cols if c not in ordered and c not in unnamed_cols]
        result_df = result_df[ordered + unnamed_cols + remaining]
        result_df.insert(0, '序号', range(1, len(result_df) + 1))

        # 8. 统计
        df_file_stats, df_customer_stats, df_shop_stats = compute_stats(data, shop_col, qty_col, customer_from_file)

        # 店铺调度统计
        total_work_minutes = (int(end_time.split(':')[0]) * 60 + int(end_time.split(':')[1])) - (start_h * 60 + start_m)
        shop_stats = data.groupby(shop_col)[qty_col].sum().reset_index()
        shop_stats.columns = [shop_col, '店铺总单量']
        shop_last_time = {}
        shop_avg_interval = {}
        for shop, items in shop_schedule_map.items():
            times = sorted([it['时间分钟'] for it in items])
            shop_last_time[shop] = times[-1] if times else 0
            shop_avg_interval[shop] = (times[-1] - times[0]) / (len(times) - 1) if len(times) > 1 else 0
        shop_stats['平均间隔(分钟)'] = shop_stats[shop_col].map(shop_avg_interval)
        shop_stats['预计完成时间'] = shop_stats[shop_col].map(
            lambda s: (base_time + timedelta(minutes=shop_last_time.get(s, 0))).strftime('%H:%M')
        )

        # 9. 生成输出文件：按日期命名保存到outputs根目录
        task_name = _next_task_name()
        output_path = os.path.join(OUTPUT_DIR, task_name)
        generate_output_excel(output_path, result_df, shop_stats, df_file_stats, df_customer_stats, df_shop_stats,
                               file_images=file_images, file_std_images=file_std_images)

        # 标记已处理：按session分组
        session_files = {}
        for item in (selected_files or []):
            session_files.setdefault(item['session_id'], []).append(item['filename'])
        for sid, fnames in session_files.items():
            _mark_processed(os.path.join(UPLOAD_DIR, sid), fnames)
        _notify_change()

        # 10. 构建预览数据
        anomaly_rows = []
        for dup in abnormal['duplicates']:
            anomaly_rows.append({'文件名': f"{dup[0]} <-> {dup[1]}", '异常原因': '内容完全相同的文件'})
        for inv in abnormal['invalid']:
            anomaly_rows.append({'文件名': inv['file'], '异常原因': inv['reason']})
        for warn in abnormal.get('warnings', []):
            anomaly_rows.append({'文件名': warn['file'], '异常原因': warn['reason']})

        summary_data = to_json_safe(result_df.head(200))

        return jsonify({
            'task_name': task_name,
            'session_id': session_id,
            'summary': summary_data,
            'summary_columns': [str(c) for c in result_df.columns],
            'total_rows': len(result_df),
            'file_stats': to_json_safe(df_file_stats),
            'customer_stats': to_json_safe(df_customer_stats),
            'shop_stats': to_json_safe(df_shop_stats),
            'shop_schedule': to_json_safe(shop_stats),
            'anomalies': anomaly_rows,
            'total_files': len(filepaths),
            'total_customers': len(df_customer_stats),
            'total_shops': len(df_shop_stats),
            'total_orders': int(data[qty_col].sum()),
        })

    except Exception as e:
        import traceback
        return jsonify({'error': str(e), 'traceback': traceback.format_exc()}), 500


@app.route('/api/download-summary/<path:filename>')
def download_summary(filename):
    """下载汇总表"""
    filepath = os.path.join(OUTPUT_DIR, filename)
    if not os.path.isfile(filepath):
        return jsonify({'error': '文件不存在'}), 404
    return send_file(filepath, as_attachment=True, download_name=filename)


@app.route('/api/summaries')
def list_summaries():
    """列出所有历史汇总表，按日期分组"""
    summaries = []
    for f in os.listdir(OUTPUT_DIR):
        if f.endswith('.xlsx'):
            fp = os.path.join(OUTPUT_DIR, f)
            summaries.append({
                'filename': f,
                'mtime': os.path.getmtime(fp),
                'size': os.path.getsize(fp)
            })
    summaries.sort(key=lambda x: x['mtime'], reverse=True)
    return jsonify({'summaries': summaries, 'count': len(summaries)})


@app.route('/api/delete-summary', methods=['POST'])
def delete_summary():
    """删除单个汇总表"""
    data = request.get_json()
    filename = data.get('filename', '')
    filepath = os.path.join(OUTPUT_DIR, filename)
    if not os.path.isfile(filepath):
        return jsonify({'error': '文件不存在'}), 404
    os.remove(filepath)
    _notify_change()
    return jsonify({'ok': True})


@app.route('/api/batch-delete-summaries', methods=['POST'])
def batch_delete_summaries():
    """批量删除汇总表"""
    data = request.get_json()
    filenames = data.get('filenames', [])
    deleted, failed = [], []
    for fn in filenames:
        fp = os.path.join(OUTPUT_DIR, fn)
        if os.path.isfile(fp):
            try:
                os.remove(fp)
                deleted.append(fn)
            except Exception:
                failed.append(fn)
        else:
            failed.append(fn)
    _notify_change()
    return jsonify({'ok': True, 'deleted': len(deleted), 'failed': failed})


if __name__ == '__main__':
    app.run(host='0.0.0.0', port=5000, debug=True)

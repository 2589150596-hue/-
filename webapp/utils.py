"""补单汇总工具 - 共享函数库（从汇总工具_v2和补单统计工具_v2提取）"""
import pandas as pd
import os
import re
import io
import difflib
import hashlib
import mimetypes
from datetime import datetime, timedelta
from openpyxl.utils import get_column_letter
from openpyxl import load_workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from collections import defaultdict
import zipfile
import xml.etree.ElementTree as ET

# ========== openpyxl webp 兼容补丁 ==========
mimetypes.add_type('image/webp', '.webp')
import openpyxl.packaging.manifest as _manifest
_original_register = _manifest.Manifest._register_mimetypes

def _patched_register_mimetypes(self, filenames):
    for fn in filenames:
        ext = os.path.splitext(fn)[-1]
        if not ext:
            continue
        mime = None
        if True in mimetypes.types_map:
            mime = mimetypes.types_map[True].get(ext) or mimetypes.types_map[True].get(ext.lower())
        else:
            mime = mimetypes.types_map.get(ext) or mimetypes.types_map.get(ext.lower())
        if mime is None:
            mime = 'application/octet-stream'
        fe = _manifest.FileExtension(ext[1:], mime)
        self.Default.append(fe)

_manifest.Manifest._register_mimetypes = _patched_register_mimetypes

# ========== 标准列名映射 ==========
COLUMN_ALIAS_MAP = [
    ('商品标题', ['商品标题', '标题/关键词', '全标题', '搜索关键词', '标题关键词', '关键词', '标题']),
    ('客户(不填)', ['客户(不填)', '客户名(店铺)', '客户名（不填）', '客户名', '客户']),
    ('店铺名称', ['店铺名称', '店铺']),
    ('商品ID', ['商品ID', 'ID', '商品编号']),
    ('主图', ['主图']),
    ('后台分享码（二维码）', ['后台分享码（二维码）', '二维码', '分享码']),
    ('下单账号', ['下单账号', '下单账户', '下单号', '下单账号(不填)', '下单(不填)']),
    ('支付账号', ['支付账号', '支付账户', '支付账号(不填)', '支付(不填)']),
    ('订单号(不填)', ['订单号(不填)', '订号(不填)', '单号(不填)', '订单号']),
    ('下单价格', ['下单价格', '价格', '单价', '下单价', '下价格']),
    ('补单数量', ['补单数量', '下单数量', '单量', '数量', '补单量']),
    ('佣金', ['佣金', '提成']),
    ('总计', ['总计', '合计', '总金额']),
    ('备注', ['备注', '消息', '说明', '附言']),
]

# ========== 文件名解析 ==========

def extract_customer_from_filename(basename):
    """从文件名 序号-客户-原名.xlsx 提取客户名"""
    name = os.path.splitext(basename)[0]
    temp = re.sub(r'^\d+[-_]', '', name)
    parts = [p.strip() for p in re.split(r'[-_]', temp) if p.strip()]
    return parts[0] if parts else ''


def parse_customer(filename):
    """从文件名解析客户名（补单统计工具版）"""
    name_without_ext = os.path.splitext(filename)[0]
    parts = name_without_ext.split("-", 2)
    if len(parts) >= 2:
        candidate = parts[1].strip()
        m = re.match(r"(\d+\.\d+)?(早补|晚补)?(.+)", candidate)
        if m:
            possible = m.group(3).strip()
            if possible:
                return possible
        return candidate
    return "未知客户"


def parse_filename(basename):
    """从文件名提取序号和店铺名"""
    name = os.path.splitext(basename)[0]
    serial_match = re.match(r'^(\d+)', name)
    serial = serial_match.group(1) if serial_match else basename
    shop_match = re.search(r'([\w一-龥]+(?:专卖店|旗舰店|专营店))', name)
    if shop_match:
        shop = shop_match.group(1)
    else:
        temp = re.sub(r'^\d+[-_]', '', name)
        temp = re.sub(r'\d{4}\.\s*\d+\.\s*\d+', '', temp)
        temp = re.sub(r'\d+\.\d+', '', temp)
        temp = re.sub(r'\d+单|\d+-\d+单', '', temp)
        temp = re.sub(r'^\d+', '', temp)
        temp = re.sub(r'\d+$', '', temp)
        temp = re.sub(r'内部|补数据|新模版', '', temp)
        temp = temp.strip('-_（）() ')
        parts = [p.strip() for p in re.split(r'[-_]', temp) if p.strip()]
        shop = parts[1] if len(parts) >= 2 else (parts[0] if len(parts) == 1 else '')
    return serial, shop


def clean_shop_name(name):
    """清理店铺名称中的日期、序号、括号等残留"""
    if not name:
        return "未知店铺"
    name = str(name).strip()
    name = re.sub(r"\.xlsx", "", name, flags=re.IGNORECASE)
    name = re.sub(r"\s*\(\d+\)\s*$", "", name)
    name = re.sub(r"-\d+-?\s*$", "", name)
    name = re.sub(r"\d+\.\d+.*$", "", name)
    name = re.sub(r"\d{4}[\.\s]+\d+[\.\s]+\d+.*$", "", name)
    return name.strip()


def normalize_shop_name(name):
    """统一店铺名格式"""
    if not name or str(name).strip() == "":
        return "未知店铺"
    name = str(name).strip()
    name = re.sub(r"[\(\)（）]", "", name)
    return name.strip()

# ========== 图片提取 ==========

def extract_wps_images(xlsx_path):
    """从WPS的xlsx文件中提取DISPIMG图片，返回 {图片ID: 二进制数据}"""
    images = {}
    try:
        with zipfile.ZipFile(xlsx_path, 'r') as z:
            if 'xl/cellimages.xml' not in z.namelist():
                return images
            rels_content = z.read('xl/_rels/cellimages.xml.rels').decode('utf-8')
            rels_root = ET.fromstring(rels_content)
            rid_to_media = {rel.get('Id'): rel.get('Target') for rel in rels_root}
            ci_content = z.read('xl/cellimages.xml').decode('utf-8')
            ci_root = ET.fromstring(ci_content)
            ns = {'etc': 'http://www.wps.cn/officeDocument/2017/etCustomData',
                  'xdr': 'http://schemas.openxmlformats.org/drawingml/2006/spreadsheetDrawing',
                  'a': 'http://schemas.openxmlformats.org/drawingml/2006/main'}
            for cellimg in ci_root.findall('etc:cellImage', ns):
                pic = cellimg.find('xdr:pic', ns)
                if pic is None:
                    continue
                cNvPr = pic.find('xdr:nvPicPr/xdr:cNvPr', ns)
                if cNvPr is None:
                    continue
                img_id = cNvPr.get('name')
                blip = pic.find('xdr:blipFill/a:blip', ns)
                if blip is None:
                    continue
                rid = blip.get('{http://schemas.openxmlformats.org/officeDocument/2006/relationships}embed')
                media_path = rid_to_media.get(rid)
                if media_path:
                    images[img_id] = z.read(f'xl/{media_path}')
    except Exception:
        pass
    return images


def extract_img_id(cell_value):
    """从 =DISPIMG(\"ID_...\",1) 中提取图片ID"""
    if isinstance(cell_value, str):
        m = re.search(r'DISPIMG\("(ID_[A-F0-9]+)"', cell_value)
        if m:
            return m.group(1)
    return None


def extract_standard_images(xlsx_path):
    """提取标准Excel中的图片，返回 {(row, col): 二进制数据}"""
    images = {}
    try:
        with zipfile.ZipFile(xlsx_path, 'r') as z:
            names = z.namelist()
            drawing_files = [n for n in names if n.startswith('xl/drawings/drawing') and n.endswith('.xml')]
            for drawing_file in drawing_files:
                rels_file = drawing_file.replace('drawings/', 'drawings/_rels/') + '.rels'
                if rels_file not in names:
                    continue
                rels_content = z.read(rels_file).decode('utf-8')
                rels_root = ET.fromstring(rels_content)
                ns_rels = {'r': 'http://schemas.openxmlformats.org/package/2006/relationships'}
                rid_to_target = {rel.get('Id'): rel.get('Target') for rel in rels_root.findall('r:Relationship', ns_rels)}
                drawing_content = z.read(drawing_file).decode('utf-8')
                drawing_root = ET.fromstring(drawing_content)
                ns = {'xdr': 'http://schemas.openxmlformats.org/drawingml/2006/spreadsheetDrawing',
                      'a': 'http://schemas.openxmlformats.org/drawingml/2006/main',
                      'r': 'http://schemas.openxmlformats.org/officeDocument/2006/relationships'}
                for anchor in drawing_root.findall('.//xdr:twoCellAnchor', ns) + drawing_root.findall('.//xdr:oneCellAnchor', ns):
                    from_elem = anchor.find('xdr:from', ns)
                    if from_elem is None:
                        continue
                    row_elem = from_elem.find('xdr:row', ns)
                    col_elem = from_elem.find('xdr:col', ns)
                    if row_elem is None or col_elem is None:
                        continue
                    row, col = int(row_elem.text), int(col_elem.text)
                    pic = anchor.find('.//xdr:pic', ns)
                    if pic is None:
                        continue
                    blip = pic.find('.//a:blip', ns)
                    if blip is None:
                        continue
                    embed = blip.get('{http://schemas.openxmlformats.org/officeDocument/2006/relationships}embed')
                    if not embed:
                        continue
                    target = rid_to_target.get(embed)
                    if not target:
                        continue
                    if target.startswith('../'):
                        target = 'xl/' + target[3:]
                    elif target.startswith('/'):
                        target = target[1:]
                    else:
                        target = 'xl/' + target
                    if target in names:
                        images[(row, col)] = z.read(target)
    except Exception:
        pass
    return images


def get_header_column_map(filepath):
    """读取Excel表头，返回 {列索引(0-based): 标准列名}"""
    try:
        wb = load_workbook(filepath, read_only=True, data_only=True)
        ws = wb.active
        mapping = {}
        for cell in ws[1]:
            if cell.value:
                col_str = str(cell.value).strip()
                for std_name, aliases in COLUMN_ALIAS_MAP:
                    if col_str in aliases or col_str == std_name:
                        mapping[cell.column - 1] = std_name
                        break
        wb.close()
        return mapping
    except Exception:
        return {}

# ========== 列名标准化 ==========

def filter_total_rows(df):
    """过滤掉数据中的'合计'汇总行"""
    if len(df) == 0:
        return df
    mask = pd.Series([True] * len(df), index=df.index)
    for col in df.columns:
        try:
            col_values = df[col].astype(str).str.strip()
            total_mask = col_values == '合计'
            if total_mask.any():
                mask &= ~total_mask
        except Exception:
            continue
    return df[mask].reset_index(drop=True)


def normalize_columns(df):
    """统一各文件的列名：精确别名 + 模糊匹配 + 内容合并"""
    rename_map = {}
    matched_originals = set()
    for std_name, aliases in COLUMN_ALIAS_MAP:
        for col in df.columns:
            if col in matched_originals:
                continue
            col_str = str(col).strip()
            if col_str in aliases:
                rename_map[col] = std_name
                matched_originals.add(col)
                break
    for col in df.columns:
        if col in matched_originals:
            continue
        col_str = str(col).strip()
        std_best_ratios = {}
        for std_name, aliases in COLUMN_ALIAS_MAP:
            candidates = [std_name] + aliases
            std_best_ratios[std_name] = max(
                difflib.SequenceMatcher(None, col_str, cand).ratio() for cand in candidates
            )
        sorted_ratios = sorted(std_best_ratios.items(), key=lambda x: x[1], reverse=True)
        best_match, best_ratio = sorted_ratios[0]
        second_ratio = sorted_ratios[1][1] if len(sorted_ratios) > 1 else 0
        if best_ratio > 0.6 and best_ratio - second_ratio > 0.15:
            rename_map[col] = best_match
            matched_originals.add(col)
    if rename_map:
        std_to_sources = {}
        for src, dst in rename_map.items():
            std_to_sources.setdefault(dst, []).append(src)
        final_rename = {}
        for dst, sources in std_to_sources.items():
            final_rename[sources[0]] = dst
            if len(sources) > 1:
                if dst not in df.columns:
                    df[dst] = pd.Series(dtype=object)
                for src in sources[1:]:
                    if src in df.columns:
                        df[dst] = df[dst].fillna(df[src])
        df.rename(columns=final_rename, inplace=True)
        to_drop = [s for dst, sources in std_to_sources.items() for s in sources[1:] if s in df.columns and s != dst]
        if to_drop:
            df.drop(columns=to_drop, inplace=True)
    if '商品标题' not in df.columns:
        for col in df.columns:
            col_str = str(col).strip()
            if col_str.startswith('Unnamed'):
                sample = df[col].dropna().astype(str)
                if len(sample) > 0:
                    avg_len = sample.str.len().mean()
                    has_dispimg = sample.str.contains('DISPIMG', na=False).any()
                    if avg_len > 20 and not has_dispimg:
                        df.rename(columns={col: '商品标题'}, inplace=True)
                        break
    return df

# ========== Excel读取 ==========

def read_excel_with_fallback(filepath):
    """读取Excel，如果找不到关键列则尝试第二行作为表头补充"""
    try:
        df = pd.read_excel(filepath, dtype={'商品ID': str})
    except ValueError:
        df = pd.read_excel(filepath, dtype=str)
    df = df.dropna(how='all')
    df = filter_total_rows(df)
    if len(df) > 5000:
        key_cols = [c for c in df.columns if not str(c).startswith('Unnamed')]
        if key_cols:
            last_idx = df[key_cols[0]].last_valid_index()
            df = df.iloc[:min(last_idx + 1 if last_idx is not None and last_idx < 5000 else 5000, 5000)]
        else:
            df = df.iloc[:5000]
    df_norm = normalize_columns(df.copy())
    has_shop = any('店铺名称' == str(c) or ('客户' in str(c) and '店铺' in str(c)) or str(c) == '客户(不填)' for c in df_norm.columns)
    has_qty = any(str(c) in ['补单数量', '单量', '数量'] for c in df_norm.columns)
    if has_shop and has_qty:
        df_norm['__drawing_row__'] = df_norm.index + 1
        return df_norm
    if len(df) >= 1:
        row2 = df.iloc[0]
        rename_map = {}
        for col in df.columns:
            col_str = str(col).strip()
            if col_str.startswith('Unnamed') or col_str == '' or col_str == 'nan':
                val = row2.get(col)
                if pd.notna(val) and str(val).strip():
                    rename_map[col] = str(val).strip()
        if rename_map:
            df_renamed = df.rename(columns=rename_map).iloc[1:].copy()
            df_renamed['__drawing_row__'] = df_renamed.index + 1
            df_renamed = df_renamed.reset_index(drop=True)
            df_renamed = normalize_columns(df_renamed)
            has_shop_r = any('店铺名称' == str(c) or ('客户' in str(c) and '店铺' in str(c)) or str(c) == '客户(不填)' for c in df_renamed.columns)
            has_qty_r = any(str(c) in ['补单数量', '单量', '数量'] for c in df_renamed.columns)
            if has_shop_r and has_qty_r:
                return df_renamed
    try:
        try:
            df2 = pd.read_excel(filepath, dtype={'商品ID': str}, header=1)
        except ValueError:
            df2 = pd.read_excel(filepath, dtype=str, header=1)
        df2 = df2.dropna(how='all')
        df2 = filter_total_rows(df2)
        if len(df2) > 5000:
            key_cols = [c for c in df2.columns if not str(c).startswith('Unnamed')]
            if key_cols:
                last_idx = df2[key_cols[0]].last_valid_index()
                df2 = df2.iloc[:min(last_idx + 1 if last_idx is not None and last_idx < 5000 else 5000, 5000)]
            else:
                df2 = df2.iloc[:5000]
        df2 = normalize_columns(df2)
        has_shop2 = any('店铺名称' == str(c) or ('客户' in str(c) and '店铺' in str(c)) or str(c) == '客户(不填)' for c in df2.columns)
        has_qty2 = any(str(c) in ['补单数量', '单量', '数量'] for c in df2.columns)
        if has_shop2 and has_qty2:
            df2['__drawing_row__'] = df2.index + 2
            return df2
    except Exception:
        pass
    df_norm['__drawing_row__'] = df_norm.index + 1
    return df_norm

# ========== 异常检测 ==========

def detect_abnormal_files(filepaths):
    """检测异常文件：重复内容、缺少关键列、无有效数据、行缺失"""
    file_hashes = {}
    duplicates = []
    invalid = []
    warnings = []
    for filepath in filepaths:
        basename = os.path.basename(filepath)
        try:
            df = read_excel_with_fallback(filepath)
            if len(df) == 0:
                invalid.append({'file': basename, 'reason': '空表（无数据行）'})
                continue
            # 找到实际的店铺列和单量列
            shop_col_name = None
            for c in df.columns:
                if '店铺名称' == str(c) or ('客户' in str(c) and '店铺' in str(c)) or str(c) == '客户(不填)':
                    shop_col_name = c
                    break
            qty_col = None
            for c in df.columns:
                if str(c) in ['补单数量', '单量', '数量']:
                    qty_col = c
                    break
            if shop_col_name is None:
                invalid.append({'file': basename, 'reason': '缺少店铺列'})
                continue
            if qty_col is None:
                invalid.append({'file': basename, 'reason': '缺少单量列'})
                continue
            # 检查列是否有实际数据
            shop_has_data = df[shop_col_name].notna().sum() > 0
            qty_has_data = df[qty_col].notna().sum() > 0
            if not shop_has_data:
                invalid.append({'file': basename, 'reason': '店铺列全为空'})
                continue
            if not qty_has_data:
                invalid.append({'file': basename, 'reason': '单量列全为空'})
                continue
            # 检查是否有行缺少下单数量或店铺名称
            # 跳过汇总行：店铺名和商品标题同时为空的视为汇总/小计行
            title_col = None
            for c in df.columns:
                if str(c) == '商品标题':
                    title_col = c
                    break
            qty_empty = df[qty_col].isna()
            shop_empty = df[shop_col_name].isna()
            title_empty = df[title_col].isna() if title_col else pd.Series([False] * len(df))
            is_summary_row = shop_empty & title_empty  # 汇总行：只有数字没有店铺和标题
            # 只统计非汇总行中缺失的数据
            qty_missing = (qty_empty & ~shop_empty & ~is_summary_row).sum()
            shop_missing = (shop_empty & ~qty_empty & ~is_summary_row).sum()
            if qty_missing > 0:
                warnings.append({'file': basename, 'reason': f'有 {qty_missing} 行缺少下单数量（已有店铺名）'})
            if shop_missing > 0:
                warnings.append({'file': basename, 'reason': f'有 {shop_missing} 行缺少店铺名称（已有数量）'})
            for c in df.columns:
                if str(c) in ['补单数量', '单量', '数量']:
                    qty_col = c
                    break
            if qty_col:
                qty_numeric = pd.to_numeric(df[qty_col], errors='coerce')
                if (qty_numeric > 0).sum() == 0:
                    invalid.append({'file': basename, 'reason': '单量列无有效数据'})
                    continue
            hash_cols = [c for c in df.columns if not str(c).startswith('Unnamed') and c != '__drawing_row__']
            subset = df[hash_cols].fillna('')
            rows = [tuple(str(v) for v in row) for _, row in subset.iterrows()]
            content_hash = hashlib.md5(str(rows).encode('utf-8')).hexdigest()
            if content_hash in file_hashes:
                duplicates.append([file_hashes[content_hash], basename])
            else:
                file_hashes[content_hash] = basename
        except Exception as e:
            invalid.append({'file': basename, 'reason': f'读取失败: {e}'})
    return {'duplicates': duplicates, 'invalid': invalid, 'warnings': warnings}

# ========== 调度算法 ==========

def assign_shops_to_brushers(shop_orders, num_workers):
    """将店铺分配给刷手：一店一刷手，各刷手总单量尽量平均"""
    if num_workers <= 1:
        return {0: shop_orders.copy()}
    shops = sorted(shop_orders.items(), key=lambda x: x[1], reverse=True)
    workers = [[] for _ in range(num_workers)]
    worker_totals = [0] * num_workers
    for shop, count in shops:
        min_total = min(worker_totals)
        candidates = [i for i, t in enumerate(worker_totals) if t == min_total]
        if len(candidates) == 1:
            best_w = candidates[0]
        else:
            best_w = candidates[0]
            best_diff = float('inf')
            for w in candidates:
                if len(workers[w]) == 0:
                    diff = count
                else:
                    avg = sum(c for _, c in workers[w]) / len(workers[w])
                    diff = abs(avg - count)
                if diff < best_diff:
                    best_diff = diff
                    best_w = w
        workers[best_w].append((shop, count))
        worker_totals[best_w] += count
    return {i: dict(w) for i, w in enumerate(workers)}


def schedule_brusher_orders(start_time_str, end_time_str, shop_orders, min_interval, brusher_name):
    """为一个刷手调度其所有店铺，保持各店铺均匀间隔，按单量降序依次错开"""
    sh, sm = map(int, start_time_str.split(':'))
    eh, em = map(int, end_time_str.split(':'))
    T = (eh * 60 + em) - (sh * 60 + sm)
    if T <= 0:
        raise ValueError("结束时间必须晚于开始时间")

    shops_sorted = sorted(shop_orders.items(), key=lambda x: x[1], reverse=True)
    total_orders = sum(count for _, count in shops_sorted)
    if total_orders == 0:
        return []

    all_orders = []
    cumulative_offset = 0.0

    for shop, count in shops_sorted:
        if count <= 0:
            continue
        interval = T / count
        times = []
        for i in range(count):
            t = cumulative_offset + i * interval
            if i > 0:
                t = max(t, times[-1] + min_interval)
            times.append(t)

        needed = (count - 1) * min_interval
        if needed > T - cumulative_offset:
            print(f"  警告: [{shop}] 需要至少 {needed} 分钟才能满足最小间隔，但剩余时段仅 {T - cumulative_offset:.0f} 分钟")

        for t in times:
            all_orders.append({'店铺': shop, '理想时间': t})

        cumulative_offset += min_interval

    all_orders.sort(key=lambda x: x['理想时间'])

    worker_orders = []
    for order in all_orders:
        shop = order['店铺']
        ideal = order['理想时间']
        same_shop_times = [t for s, t in worker_orders if s == shop]
        last_same = max(same_shop_times) if same_shop_times else -float('inf')
        earliest = max(ideal, last_same + min_interval)
        if earliest > T:
            earliest = T
            if last_same + min_interval > T:
                continue
        worker_orders.append((shop, earliest))

    return [{'店铺': s, '时间分钟': round(t, 1), '刷手': brusher_name} for s, t in worker_orders]


def auto_schedule(start_time_str, end_time_str, shop_orders, num_workers, min_interval):
    """自动调度：先分配刷手，再各自调度"""
    assignments = assign_shops_to_brushers(shop_orders, num_workers)
    all_results = []
    for w_idx, brusher_shops in assignments.items():
        brusher_name = f'刷手{w_idx + 1}'
        results = schedule_brusher_orders(start_time_str, end_time_str, brusher_shops, min_interval, brusher_name)
        all_results.extend(results)
    all_results.sort(key=lambda x: (x['时间分钟'], x['店铺']))
    return all_results


def interleave_shop_links(data, shop_col):
    """同一店铺内不同链接（商品标题）按轮询交叉排列"""
    link_col = None
    for col in data.columns:
        if str(col) == '商品标题':
            link_col = col
            break
    if link_col is None:
        return data
    result_rows = []
    for shop, group in data.groupby(shop_col, sort=False):
        links = {}
        for _, row in group.iterrows():
            title = str(row.get(link_col, ''))
            links.setdefault(title, []).append(row)
        link_lists = list(links.values())
        max_len = max(len(lst) for lst in link_lists)
        interleaved = []
        for i in range(max_len):
            for lst in link_lists:
                if i < len(lst):
                    interleaved.append(lst[i])
        result_rows.extend(interleaved)
    return pd.DataFrame(result_rows).reset_index(drop=True)

# ========== 统计计算 ==========

def compute_stats(data, shop_col, qty_col, customer_from_file):
    """计算文件统计、客户统计、店铺统计"""
    file_stats = []
    customer_stats = {}
    shop_stats_summary = {}
    for basename in sorted(customer_from_file.keys()):
        customer = customer_from_file[basename]
        file_data = data[data['来源文件'] == basename]
        qty_sum = file_data[qty_col].sum()
        file_stats.append({
            '文件名': basename,
            '客户名': customer,
            '补单数量': int(qty_sum) if qty_sum == int(qty_sum) else qty_sum
        })
        customer_stats[customer] = customer_stats.get(customer, 0) + qty_sum
        for shop_name, group in file_data.groupby(shop_col):
            shop_stats_summary[shop_name] = shop_stats_summary.get(shop_name, 0) + group[qty_col].sum()
    df_file_stats = pd.DataFrame(file_stats)
    df_customer_stats = pd.DataFrame(
        [{'客户名': k, '补单数量': int(v) if v == int(v) else v}
         for k, v in sorted(customer_stats.items(), key=lambda x: x[1], reverse=True)]
    )
    df_shop_stats = pd.DataFrame(
        [{'店铺名': k, '补单数量': int(v) if v == int(v) else v}
         for k, v in sorted(shop_stats_summary.items(), key=lambda x: x[1], reverse=True)]
    )
    return df_file_stats, df_customer_stats, df_shop_stats

# ========== Excel 导出 ==========

def _embed_image_in_cell(ws, img, col_idx, row_idx):
    """将图片锚定到单元格内并居中偏移"""
    try:
        col_letter = get_column_letter(col_idx)
        col_w = ws.column_dimensions.get(col_letter)
        col_w_ch = col_w.width if col_w and col_w.width else 14
        row_h_pt = ws.row_dimensions[row_idx].height if ws.row_dimensions[row_idx].height else 65
        offset_x = int(max(0, (col_w_ch * 7.5 - img.width) / 2) * 12700)
        offset_y = int(max(0, (row_h_pt * 1.33 - img.height) / 2) * 12700)

        from openpyxl.drawing.spreadsheet_drawing import AnchorMarker, TwoCellAnchor
        from_m = AnchorMarker(col=col_idx - 1, colOff=offset_x, row=row_idx - 1, rowOff=offset_y)
        to_m = AnchorMarker(col=col_idx, colOff=0, row=row_idx, rowOff=0)
        img.anchor = TwoCellAnchor('oneCell', _from=from_m, to=to_m)
        ws.add_image(img)
    except Exception:
        # 图片嵌入失败时使用简单锚定作为后备
        try:
            cell_addr = f"{col_letter}{row_idx}"
            ws.add_image(img, cell_addr)
        except Exception:
            pass


def generate_output_excel(output_path, result_df, shop_stats, df_file_stats, df_customer_stats, df_shop_stats,
                          file_images=None, file_std_images=None):
    """生成最终的汇总Excel文件，含图片插入和格式美化"""
    file_images = file_images or {}
    file_std_images = file_std_images or {}

    output_df = result_df.drop(columns=['__drawing_row__'], errors='ignore')
    with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
        output_df.to_excel(writer, sheet_name='汇总表', index=False)
        shop_stats.to_excel(writer, sheet_name='店铺统计', index=False)
        df_file_stats.to_excel(writer, sheet_name='按文件统计', index=False)
        df_customer_stats.to_excel(writer, sheet_name='按客户统计', index=False)
        df_shop_stats.to_excel(writer, sheet_name='按店铺统计', index=False)

    wb = load_workbook(output_path)
    ws = wb['汇总表']
    col_idx_map = {cell.value: cell.column for cell in ws[1]}

    header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    header_font = Font(name='微软雅黑', size=11, bold=True, color='FFFFFF')
    header_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    even_fill = PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid')
    odd_fill = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')
    thin_border = Border(
        left=Side(style='thin', color='D9D9D9'),
        right=Side(style='thin', color='D9D9D9'),
        top=Side(style='thin', color='D9D9D9'),
        bottom=Side(style='thin', color='D9D9D9')
    )
    center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    left_align = Alignment(horizontal='left', vertical='center', wrap_text=True)
    right_align = Alignment(horizontal='right', vertical='center')

    # 图片列宽度
    pic_cols = ['主图', '后台分享码（二维码）']
    for pic_col in pic_cols:
        if pic_col in col_idx_map:
            ws.column_dimensions[get_column_letter(col_idx_map[pic_col])].width = 14
    ws.row_dimensions[1].height = 30

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_align
        cell.border = thin_border

    # ========== 插入图片 ==========
    inserted = 0
    for excel_row_idx, (df_idx, row_series) in enumerate(result_df.iterrows(), start=2):
        row_fill = even_fill if (excel_row_idx % 2 == 0) else odd_fill
        ws.row_dimensions[excel_row_idx].height = 65

        for col_idx in range(1, ws.max_column + 1):
            cell = ws.cell(row=excel_row_idx, column=col_idx)
            cell.border = thin_border
            cell.fill = row_fill
            cell.alignment = left_align

        source_file = row_series.get('来源文件')
        if not source_file:
            continue

        orig_path = None
        for fp in file_images:
            if os.path.basename(fp) == source_file:
                orig_path = fp
                break
        if not orig_path:
            continue

        images = file_images.get(orig_path, {})
        std_images = file_std_images.get(orig_path, {})
        drawing_row = row_series.get('__drawing_row__')

        for col_name in result_df.columns:
            if col_name in ('序号', '执行时间', '距离首单(分钟)', '店铺总单量', '来源文件'):
                continue
            cell_value = row_series[col_name]
            img_id = extract_img_id(cell_value)
            col_idx = col_idx_map.get(col_name)
            if not col_idx:
                continue

            cell = ws.cell(row=excel_row_idx, column=col_idx)
            has_img = False

            # WPS图片
            if img_id and img_id in images:
                cell.value = None
                try:
                    img = XLImage(io.BytesIO(images[img_id]))
                    img.width = 70
                    img.height = 70
                    _embed_image_in_cell(ws, img, col_idx, excel_row_idx)
                    inserted += 1
                    has_img = True
                except Exception as e:
                    print(f"  图片插入警告: {e}")
            elif isinstance(cell_value, str) and 'DISPIMG' in cell_value:
                cell.value = None
                has_img = True

            # 标准Excel图片
            if not has_img and col_name in std_images and drawing_row is not None:
                dr = int(drawing_row) if drawing_row is not None else None
                if dr is not None and dr in std_images[col_name]:
                    img_data = std_images[col_name][dr]
                    cell.value = None
                    try:
                        img = XLImage(io.BytesIO(img_data))
                        img.width = 70
                        img.height = 70
                        _embed_image_in_cell(ws, img, col_idx, excel_row_idx)
                        inserted += 1
                    except Exception as e:
                        print(f"  图片插入警告: {e}")

    # ========== 列宽和对齐 ==========
    col_widths = {
        '序号': 6, '执行时间': 12, '刷手': 10, '距离首单(分钟)': 14, '店铺总单量': 12,
        '商品标题': 50, '店铺名称': 35, '客户(不填)': 15, '下单账号': 15,
        '支付账号': 15, '下单价格': 12, '补单数量': 12, '佣金': 10, '总计': 10,
        '来源文件': 25, '商品ID': 18, '订单号(不填)': 15, '备注': 30,
    }
    for col_name, width in col_widths.items():
        if col_name in col_idx_map:
            ws.column_dimensions[get_column_letter(col_idx_map[col_name])].width = width

    # 数字列右对齐
    num_cols = ['序号', '距离首单(分钟)', '店铺总单量', '下单价格', '补单数量', '佣金', '总计']
    for col_name in num_cols:
        if col_name in col_idx_map:
            c = col_idx_map[col_name]
            for row in range(1, ws.max_row + 1):
                ws.cell(row=row, column=c).alignment = center_align if row == 1 else right_align

    # 时间列居中
    for col_name in ['执行时间']:
        if col_name in col_idx_map:
            c = col_idx_map[col_name]
            for row in range(1, ws.max_row + 1):
                ws.cell(row=row, column=c).alignment = center_align

    # 商品ID文本格式
    if '商品ID' in col_idx_map:
        c = col_idx_map['商品ID']
        for row in range(2, ws.max_row + 1):
            cell = ws.cell(row=row, column=c)
            if cell.value is not None:
                cell.number_format = '@'

    ws.freeze_panes = 'A2'
    wb.save(output_path)
    return output_path

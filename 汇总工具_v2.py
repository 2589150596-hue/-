import pandas as pd
import os
import glob
import zipfile
import xml.etree.ElementTree as ET
import re
import io
import difflib
import hashlib
import shutil
import mimetypes
from datetime import datetime, timedelta
from openpyxl.utils import get_column_letter
from openpyxl import load_workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side

# 修复openpyxl保存时无法识别webp等图片格式的问题
# 方案1: 注册webp到mimetypes
mimetypes.add_type('image/webp', '.webp')

# 方案2: 给openpyxl的manifest打补丁，遇到未知扩展名时不崩溃
import openpyxl.packaging.manifest as _manifest
_original_register = _manifest.Manifest._register_mimetypes

def _patched_register_mimetypes(self, filenames):
    for fn in filenames:
        ext = os.path.splitext(fn)[-1]
        if not ext:
            continue
        # 兼容不同Python版本的mimetypes API
        mime = None
        if True in mimetypes.types_map:
            mime = mimetypes.types_map[True].get(ext) or mimetypes.types_map[True].get(ext.lower())
        else:
            mime = mimetypes.types_map.get(ext) or mimetypes.types_map.get(ext.lower())
        if mime is None:
            mime = 'application/octet-stream'
        fe = _manifest.FileExtension(ext[1:], mime)
        self.Default.append(fe)

_manifest.Manifest._register_mimetypes = _patched_register_mimetypes

# 标准列名及其别名列表（按优先级排序）
COLUMN_ALIAS_MAP = [
    ('商品标题', ['商品标题', '标题/关键词', '全标题', '搜索关键词', '标题关键词', '关键词', '标题']),
    ('客户(不填)', ['客户(不填)', '客户名(店铺)', '客户名（不填）', '客户名', '客户']),
    ('店铺名称', ['店铺名称', '店铺']),
    ('商品ID', ['商品ID', 'ID', '商品编号']),
    ('主图', ['主图']),
    ('后台分享码（二维码）', ['后台分享码（二维码）', '二维码', '分享码']),
    ('下单账号', ['下单账号', '下单账户', '下单号', '下单账号(不填)', '下单(不填)']),
    ('支付账号', ['支付账号', '支付账户', '支付账号(不填)', '支付(不填)']),
    ('订单号(不填)', ['订单号(不填)', '订号(不填)', '单号(不填)', '订单号']),
    ('下单价格', ['下单价格', '价格', '单价', '下单价', '下价格']),
    ('补单数量', ['补单数量', '下单数量', '单量', '数量', '补单量']),
    ('佣金', ['佣金', '提成']),
    ('总计', ['总计', '合计', '总金额']),
    ('备注', ['备注', '消息', '说明', '附言']),
]


def extract_customer_from_filename(basename):
    """从文件名 序号-客户-原名.xlsx 提取客户名"""
    name = os.path.splitext(basename)[0]
    temp = re.sub(r'^\d+[-_]', '', name)
    parts = [p.strip() for p in re.split(r'[-_]', temp) if p.strip()]
    return parts[0] if parts else ''


def parse_filename(basename):
    """从文件名提取序号和店铺名"""
    name = os.path.splitext(basename)[0]
    serial_match = re.match(r'^(\d+)', name)
    serial = serial_match.group(1) if serial_match else basename

    shop_match = re.search(r'([\w一-龥]+(?:专卖店|旗舰店|专营店))', name)
    if shop_match:
        shop = shop_match.group(1)
    else:
        temp = re.sub(r'^\d+[-_]', '', name)
        temp = re.sub(r'\d{4}\.\s*\d+\.\s*\d+', '', temp)
        temp = re.sub(r'\d+\.\d+', '', temp)
        temp = re.sub(r'\d+单|\d+-\d+单', '', temp)
        temp = re.sub(r'^\d+', '', temp)
        temp = re.sub(r'\d+$', '', temp)
        temp = re.sub(r'内部|补数据|新模版', '', temp)
        temp = temp.strip('-_（）() ')
        parts = [p.strip() for p in re.split(r'[-_]', temp) if p.strip()]
        if len(parts) >= 2:
            shop = parts[1]
        elif len(parts) == 1:
            shop = parts[0]
        else:
            shop = ''
    return serial, shop


def extract_wps_images(xlsx_path):
    """从WPS的xlsx文件中提取DISPIMG图片，返回 {图片ID: 二进制数据}"""
    images = {}
    try:
        with zipfile.ZipFile(xlsx_path, 'r') as z:
            if 'xl/cellimages.xml' not in z.namelist():
                return images

            rels_content = z.read('xl/_rels/cellimages.xml.rels').decode('utf-8')
            rels_root = ET.fromstring(rels_content)
            rid_to_media = {rel.get('Id'): rel.get('Target') for rel in rels_root}

            ci_content = z.read('xl/cellimages.xml').decode('utf-8')
            ci_root = ET.fromstring(ci_content)
            ns = {'etc': 'http://www.wps.cn/officeDocument/2017/etCustomData',
                  'xdr': 'http://schemas.openxmlformats.org/drawingml/2006/spreadsheetDrawing',
                  'a': 'http://schemas.openxmlformats.org/drawingml/2006/main'}

            for cellimg in ci_root.findall('etc:cellImage', ns):
                pic = cellimg.find('xdr:pic', ns)
                if pic is None:
                    continue
                cNvPr = pic.find('xdr:nvPicPr/xdr:cNvPr', ns)
                if cNvPr is None:
                    continue
                img_id = cNvPr.get('name')
                blip = pic.find('xdr:blipFill/a:blip', ns)
                if blip is None:
                    continue
                rid = blip.get('{http://schemas.openxmlformats.org/officeDocument/2006/relationships}embed')
                media_path = rid_to_media.get(rid)
                if media_path:
                    images[img_id] = z.read(f'xl/{media_path}')
    except Exception as e:
        print(f"  提取图片失败: {e}")
    return images


def extract_img_id(cell_value):
    """从 =DISPIMG(\"ID_...\",1) 中提取图片ID"""
    if isinstance(cell_value, str):
        m = re.search(r'DISPIMG\("(ID_[A-F0-9]+)"', cell_value)
        if m:
            return m.group(1)
    return None


def filter_total_rows(df):
    """过滤掉数据中的'合计'汇总行"""
    if len(df) == 0:
        return df
    mask = pd.Series([True] * len(df), index=df.index)
    for col in df.columns:
        try:
            col_values = df[col].astype(str).str.strip()
            total_mask = col_values == '合计'
            if total_mask.any():
                mask &= ~total_mask
        except Exception:
            continue
    return df[mask].reset_index(drop=True)


def extract_standard_images(xlsx_path):
    """提取标准Excel中的图片，返回 {(row, col): 二进制数据}"""
    images = {}
    try:
        with zipfile.ZipFile(xlsx_path, 'r') as z:
            names = z.namelist()
            drawing_files = [n for n in names if n.startswith('xl/drawings/drawing') and n.endswith('.xml')]
            for drawing_file in drawing_files:
                rels_file = drawing_file.replace('drawings/', 'drawings/_rels/') + '.rels'
                if rels_file not in names:
                    continue
                rels_content = z.read(rels_file).decode('utf-8')
                rels_root = ET.fromstring(rels_content)
                ns_rels = {'r': 'http://schemas.openxmlformats.org/package/2006/relationships'}
                rid_to_target = {rel.get('Id'): rel.get('Target') for rel in rels_root.findall('r:Relationship', ns_rels)}
                drawing_content = z.read(drawing_file).decode('utf-8')
                drawing_root = ET.fromstring(drawing_content)
                ns = {
                    'xdr': 'http://schemas.openxmlformats.org/drawingml/2006/spreadsheetDrawing',
                    'a': 'http://schemas.openxmlformats.org/drawingml/2006/main',
                    'r': 'http://schemas.openxmlformats.org/officeDocument/2006/relationships'
                }
                for anchor in drawing_root.findall('.//xdr:twoCellAnchor', ns) + drawing_root.findall('.//xdr:oneCellAnchor', ns):
                    from_elem = anchor.find('xdr:from', ns)
                    if from_elem is None:
                        continue
                    row_elem = from_elem.find('xdr:row', ns)
                    col_elem = from_elem.find('xdr:col', ns)
                    if row_elem is None or col_elem is None:
                        continue
                    row = int(row_elem.text)
                    col = int(col_elem.text)
                    pic = anchor.find('.//xdr:pic', ns)
                    if pic is None:
                        continue
                    blip = pic.find('.//a:blip', ns)
                    if blip is None:
                        continue
                    embed = blip.get('{http://schemas.openxmlformats.org/officeDocument/2006/relationships}embed')
                    if not embed:
                        continue
                    target = rid_to_target.get(embed)
                    if not target:
                        continue
                    if target.startswith('../'):
                        target = 'xl/' + target[3:]
                    elif target.startswith('/'):
                        target = target[1:]
                    else:
                        target = 'xl/' + target
                    if target in names:
                        images[(row, col)] = z.read(target)
    except Exception as e:
        print(f"  提取标准图片失败: {e}")
    return images


def get_header_column_map(filepath):
    """读取Excel表头，返回 {列索引(0-based): 标准列名}"""
    try:
        wb = load_workbook(filepath, read_only=True, data_only=True)
        ws = wb.active
        mapping = {}
        for cell in ws[1]:
            if cell.value:
                col_str = str(cell.value).strip()
                for std_name, aliases in COLUMN_ALIAS_MAP:
                    if col_str in aliases or col_str == std_name:
                        mapping[cell.column - 1] = std_name
                        break
        wb.close()
        return mapping
    except Exception as e:
        print(f"  读取表头失败: {e}")
        return {}


def normalize_columns(df):
    """统一各文件的列名：精确别名 + 模糊匹配 + 内容合并"""
    rename_map = {}
    matched_originals = set()

    for std_name, aliases in COLUMN_ALIAS_MAP:
        for col in df.columns:
            if col in matched_originals:
                continue
            col_str = str(col).strip()
            if col_str in aliases:
                rename_map[col] = std_name
                matched_originals.add(col)
                break

    for col in df.columns:
        if col in matched_originals:
            continue
        col_str = str(col).strip()
        std_best_ratios = {}
        for std_name, aliases in COLUMN_ALIAS_MAP:
            candidates = [std_name] + aliases
            std_best_ratios[std_name] = max(
                difflib.SequenceMatcher(None, col_str, cand).ratio() for cand in candidates
            )
        sorted_ratios = sorted(std_best_ratios.items(), key=lambda x: x[1], reverse=True)
        best_match, best_ratio = sorted_ratios[0]
        second_ratio = sorted_ratios[1][1] if len(sorted_ratios) > 1 else 0
        if best_ratio > 0.6 and best_ratio - second_ratio > 0.15:
            rename_map[col] = best_match
            matched_originals.add(col)

    if rename_map:
        std_to_sources = {}
        for src, dst in rename_map.items():
            std_to_sources.setdefault(dst, []).append(src)

        final_rename = {}
        for dst, sources in std_to_sources.items():
            final_rename[sources[0]] = dst
            if len(sources) > 1:
                if dst not in df.columns:
                    df[dst] = pd.Series(dtype=object)
                for src in sources[1:]:
                    if src in df.columns:
                        df[dst] = df[dst].fillna(df[src])

        df.rename(columns=final_rename, inplace=True)
        to_drop = [s for dst, sources in std_to_sources.items() for s in sources[1:] if s in df.columns and s != dst]
        if to_drop:
            df.drop(columns=to_drop, inplace=True)

    if '商品标题' not in df.columns:
        for col in df.columns:
            col_str = str(col).strip()
            if col_str.startswith('Unnamed'):
                sample = df[col].dropna().astype(str)
                if len(sample) > 0:
                    avg_len = sample.str.len().mean()
                    has_dispimg = sample.str.contains('DISPIMG', na=False).any()
                    if avg_len > 20 and not has_dispimg:
                        df.rename(columns={col: '商品标题'}, inplace=True)
                        break

    return df


def read_excel_with_fallback(filepath):
    """读取Excel，如果找不到关键列则尝试第二行作为表头补充"""
    try:
        df = pd.read_excel(filepath, dtype={'商品ID': str})
    except ValueError:
        df = pd.read_excel(filepath, dtype=str)
    df = df.dropna(how='all')
    df = filter_total_rows(df)

    if len(df) > 5000:
        key_cols = [c for c in df.columns if not str(c).startswith('Unnamed')]
        if key_cols:
            last_idx = df[key_cols[0]].last_valid_index()
            if last_idx is not None and last_idx < 5000:
                df = df.iloc[:last_idx + 1]
            else:
                df = df.iloc[:5000]
        else:
            df = df.iloc[:5000]

    df_norm = normalize_columns(df.copy())

    has_shop = any('店铺名称' == str(c) or ('客户' in str(c) and '店铺' in str(c)) or str(c) == '客户(不填)' for c in df_norm.columns)
    has_qty = any(str(c) in ['补单数量', '单量', '数量'] for c in df_norm.columns)

    if has_shop and has_qty:
        df_norm['__drawing_row__'] = df_norm.index + 1
        return df_norm

    # 尝试用第二行补充列名（某些表格列名分布在两行）
    if len(df) >= 1:
        row2 = df.iloc[0]
        rename_map = {}
        for col in df.columns:
            col_str = str(col).strip()
            if col_str.startswith('Unnamed') or col_str == '' or col_str == 'nan':
                val = row2.get(col)
                if pd.notna(val) and str(val).strip():
                    rename_map[col] = str(val).strip()
        if rename_map:
            df_renamed = df.rename(columns=rename_map).iloc[1:].copy()
            df_renamed['__drawing_row__'] = df_renamed.index + 1
            df_renamed = df_renamed.reset_index(drop=True)
            df_renamed = normalize_columns(df_renamed)
            has_shop_r = any('店铺名称' == str(c) or ('客户' in str(c) and '店铺' in str(c)) or str(c) == '客户(不填)' for c in df_renamed.columns)
            has_qty_r = any(str(c) in ['补单数量', '单量', '数量'] for c in df_renamed.columns)
            if has_shop_r and has_qty_r:
                return df_renamed

    # 最后尝试直接用第二行作为表头
    try:
        try:
            df2 = pd.read_excel(filepath, dtype={'商品ID': str}, header=1)
        except ValueError:
            df2 = pd.read_excel(filepath, dtype=str, header=1)
        df2 = df2.dropna(how='all')
        df2 = filter_total_rows(df2)
        if len(df2) > 5000:
            key_cols = [c for c in df2.columns if not str(c).startswith('Unnamed')]
            if key_cols:
                last_idx = df2[key_cols[0]].last_valid_index()
                if last_idx is not None and last_idx < 5000:
                    df2 = df2.iloc[:last_idx + 1]
                else:
                    df2 = df2.iloc[:5000]
            else:
                df2 = df2.iloc[:5000]
        df2 = normalize_columns(df2)
        has_shop2 = any('店铺名称' == str(c) or ('客户' in str(c) and '店铺' in str(c)) or str(c) == '客户(不填)' for c in df2.columns)
        has_qty2 = any(str(c) in ['补单数量', '单量', '数量'] for c in df2.columns)
        if has_shop2 and has_qty2:
            df2['__drawing_row__'] = df2.index + 2
            return df2
    except Exception:
        pass

    df_norm['__drawing_row__'] = df_norm.index + 1
    return df_norm


def apply_customer_names(files):
    """
    将文件名中的客户名写入原始Excel的'客户(不填)'列（仅当该列为空时）。
    操作前自动备份为 .backup.xlsx。
    """
    for filepath in files:
        basename = os.path.basename(filepath)
        customer = extract_customer_from_filename(basename)
        if not customer:
            continue

        try:
            df = read_excel_with_fallback(filepath)

            if '客户(不填)' not in df.columns:
                continue

            empty_count = df['客户(不填)'].isna().sum() + (df['客户(不填)'] == '').sum()
            if empty_count == 0:
                continue

            backup_path = filepath + '.backup.xlsx'
            if not os.path.exists(backup_path):
                shutil.copy2(filepath, backup_path)

            wb = load_workbook(filepath)
            ws = wb.active

            customer_col = None
            for cell in ws[1]:
                if cell.value and str(cell.value).strip() == '客户(不填)':
                    customer_col = cell.column
                    break

            if customer_col is None:
                continue

            for row in range(2, ws.max_row + 1):
                try:
                    cell = ws.cell(row=row, column=customer_col)
                    if cell.value is None or str(cell.value).strip() == '':
                        cell.value = customer
                except AttributeError:
                    # 跳过合并单元格
                    pass

            wb.save(filepath)
            print(f"  [{basename}] 已自动填充客户名: {customer}")
        except Exception as e:
            print(f"  [{basename}] 填充客户名失败: {e}")


def detect_abnormal_files(files):
    """
    检测异常文件：
    - 内容完全相同的文件
    - 无法读取或缺少关键列的文件
    - 有效数据为0的文件
    返回: {'duplicates': [[file1, file2], ...], 'invalid': [{'file': ..., 'reason': ...}, ...]}
    """
    file_hashes = {}
    duplicates = []
    invalid = []

    for filepath in files:
        basename = os.path.basename(filepath)
        try:
            df = read_excel_with_fallback(filepath)

            if len(df) == 0:
                invalid.append({'file': basename, 'reason': '空表（无数据行）'})
                continue

            has_shop = any('店铺名称' == str(c) or ('客户' in str(c) and '店铺' in str(c)) or str(c) == '客户(不填)' for c in df.columns)
            has_qty = any(str(c) in ['补单数量', '单量', '数量'] for c in df.columns)

            if not has_shop:
                invalid.append({'file': basename, 'reason': '缺少店铺列'})
                continue
            if not has_qty:
                invalid.append({'file': basename, 'reason': '缺少单量列'})
                continue

            qty_col = None
            for c in df.columns:
                if str(c) in ['补单数量', '单量', '数量']:
                    qty_col = c
                    break

            if qty_col:
                qty_numeric = pd.to_numeric(df[qty_col], errors='coerce')
                valid_count = (qty_numeric > 0).sum()
                if valid_count == 0:
                    invalid.append({'file': basename, 'reason': '单量列无有效数据（全0或空）'})
                    continue

            hash_cols = [c for c in df.columns if not str(c).startswith('Unnamed') and c != '__drawing_row__']
            subset = df[hash_cols].fillna('')
            rows = [tuple(str(v) for v in row) for _, row in subset.iterrows()]
            content_hash = hashlib.md5(str(rows).encode('utf-8')).hexdigest()

            if content_hash in file_hashes:
                duplicates.append([file_hashes[content_hash], basename])
            else:
                file_hashes[content_hash] = basename

        except Exception as e:
            invalid.append({'file': basename, 'reason': f'读取失败: {e}'})

    return {'duplicates': duplicates, 'invalid': invalid}


def assign_shops_to_brushers(shop_orders, num_workers):
    """
    将店铺分配给刷手。
    目标：1) 每个店铺只给一个刷手 2) 各刷手总单量尽量平均 3) 单量相近的店铺尽量聚类
    返回: {brusher_idx: {shop: count}}
    """
    if num_workers <= 1:
        return {0: shop_orders.copy()}

    shops = sorted(shop_orders.items(), key=lambda x: x[1], reverse=True)
    workers = [[] for _ in range(num_workers)]
    worker_totals = [0] * num_workers

    for shop, count in shops:
        min_total = min(worker_totals)
        candidates = [i for i, t in enumerate(worker_totals) if t == min_total]

        if len(candidates) == 1:
            best_w = candidates[0]
        else:
            best_w = candidates[0]
            best_diff = float('inf')
            for w in candidates:
                if len(workers[w]) == 0:
                    diff = count
                else:
                    avg = sum(c for _, c in workers[w]) / len(workers[w])
                    diff = abs(avg - count)
                if diff < best_diff:
                    best_diff = diff
                    best_w = w

        workers[best_w].append((shop, count))
        worker_totals[best_w] += count

    return {i: dict(w) for i, w in enumerate(workers)}


def interleave_shop_links(data, shop_col):
    """同一店铺内不同链接（商品标题）按轮询交叉排列，避免同一链接连续下单"""
    link_col = None
    for col in data.columns:
        if str(col) == '商品标题':
            link_col = col
            break
    if link_col is None:
        return data

    result_rows = []
    for shop, group in data.groupby(shop_col, sort=False):
        links = {}
        for _, row in group.iterrows():
            title = str(row.get(link_col, ''))
            links.setdefault(title, []).append(row)
        link_lists = list(links.values())
        max_len = max(len(lst) for lst in link_lists)
        interleaved = []
        for i in range(max_len):
            for lst in link_lists:
                if i < len(lst):
                    interleaved.append(lst[i])
        result_rows.extend(interleaved)

    return pd.DataFrame(result_rows).reset_index(drop=True)


def schedule_brusher_orders(start_time_str, end_time_str, shop_orders, min_interval, brusher_name):
    """为一个刷手调度其所有店铺，按工作时间平均分配间隔，小单量店铺自动错开首单时间"""
    sh, sm = map(int, start_time_str.split(':'))
    eh, em = map(int, end_time_str.split(':'))
    T = (eh * 60 + em) - (sh * 60 + sm)
    if T <= 0:
        raise ValueError("结束时间必须晚于开始时间")

    # 按单量降序排列店铺，计算起始偏移（单量越多越先排，单量少的往后错开）
    sorted_shops = sorted(shop_orders.items(), key=lambda x: x[1], reverse=True)
    num_shops = len(sorted_shops)
    stagger_step = min_interval / max(num_shops, 1)
    shop_offset = {}
    for rank, (shop, count) in enumerate(sorted_shops):
        shop_offset[shop] = rank * stagger_step

    all_orders = []
    for shop, count in shop_orders.items():
        if count <= 0:
            continue
        interval = T / count
        offset = shop_offset.get(shop, 0)
        times = []
        for i in range(count):
            t = offset + i * interval
            if i > 0:
                t = max(t, times[-1] + min_interval)
            times.append(t)

        needed = (count - 1) * min_interval
        if needed > T:
            print(f"  警告: [{shop}] 需要至少 {needed} 分钟才能满足最小间隔，但工作时段仅 {T} 分钟")

        for t in times:
            all_orders.append({'店铺': shop, '理想时间': t})

    all_orders.sort(key=lambda x: x['理想时间'])

    worker_orders = []
    for order in all_orders:
        shop = order['店铺']
        ideal = order['理想时间']

        same_shop_times = [t for s, t in worker_orders if s == shop]
        last_same = max(same_shop_times) if same_shop_times else -float('inf')
        earliest = max(ideal, last_same + min_interval)

        if earliest > T:
            earliest = T
            if last_same + min_interval > T:
                continue

        worker_orders.append((shop, earliest))

    return [{'店铺': s, '时间分钟': round(t, 1), '刷手': brusher_name} for s, t in worker_orders]


def auto_schedule(start_time_str, end_time_str, shop_orders, num_workers, min_interval):
    """
    自动调度：先按店铺分配给刷手（一店一刷手），再为每个刷手内部调度。
    返回: list[{'店铺':..., '时间分钟':..., '刷手':...}]
    """
    assignments = assign_shops_to_brushers(shop_orders, num_workers)
    all_results = []

    for w_idx, brusher_shops in assignments.items():
        brusher_name = f'刷手{w_idx + 1}'
        results = schedule_brusher_orders(start_time_str, end_time_str, brusher_shops, min_interval, brusher_name)
        all_results.extend(results)

    all_results.sort(key=lambda x: (x['时间分钟'], x['店铺']))
    return all_results


def process_folder(folder_path, output_path=None, start_time_str="08:00", end_time_str="23:00",
                   num_workers=1, min_interval=20, apply_customer=True):
    patterns = [os.path.join(folder_path, "*.xlsx"), os.path.join(folder_path, "*.xls")]
    files = []
    for p in patterns:
        files.extend(glob.glob(p))
    files = [f for f in files if not os.path.basename(f).startswith('~$') and not f.endswith('.backup.xlsx')]

    if not files:
        print(f"未在 {folder_path} 找到Excel文件")
        return

    print(f"找到 {len(files)} 个文件")

    # === 异常检测 ===
    print("\n正在检测异常文件...")
    abnormal = detect_abnormal_files(files)

    has_issues = False
    if abnormal['duplicates']:
        has_issues = True
        print("\n[警告] 发现内容完全相同的文件：")
        for dup in abnormal['duplicates']:
            print(f"   {dup[0]}  <->  {dup[1]}")

    if abnormal['invalid']:
        has_issues = True
        print("\n[警告] 发现异常文件：")
        for inv in abnormal['invalid']:
            print(f"   [{inv['file']}] {inv['reason']}")

    if has_issues:
        print("\n以上异常文件将跳过处理。")

    invalid_basenames = set()
    for inv in abnormal['invalid']:
        invalid_basenames.add(inv['file'])
    for dup in abnormal['duplicates']:
        invalid_basenames.update(dup)

    valid_files = [f for f in files if os.path.basename(f) not in invalid_basenames]

    if not valid_files:
        print("没有可处理的正常文件")
        return

    # === 自动填充客户名（仅在内存中处理，不修改原文件，避免破坏WPS图片格式） ===
    # 原文件的 openpyxl 保存会丢失 WPS 的 cellimages.xml，导致图片提取失败
    # 实际填充逻辑已在下方读取循环中实现
    if apply_customer:
        print("\n将自动从文件名提取客户名填充到汇总表...")

    # === 读取数据 ===
    all_data = []
    file_images = {}
    file_std_images = {}
    customer_from_file = {}

    for f in valid_files:
        basename = os.path.basename(f)
        customer = extract_customer_from_filename(basename)
        customer_from_file[basename] = customer

        try:
            images = extract_wps_images(f)
            file_images[f] = images
            print(f"  [{basename}] 提取了 {len(images)} 张WPS图片")

            # 提取标准Excel图片
            std_images_raw = extract_standard_images(f)
            std_images_mapped = {}
            if std_images_raw:
                header_map = get_header_column_map(f)
                for (row, col), img_data in std_images_raw.items():
                    std_col = header_map.get(col)
                    if std_col:
                        std_images_mapped.setdefault(std_col, {})[row] = img_data
                file_std_images[f] = std_images_mapped
                total_std = sum(len(v) for v in std_images_mapped.values())
                if total_std > 0:
                    print(f"  [{basename}] 提取了 {total_std} 张标准图片")
            else:
                file_std_images[f] = {}

            df = read_excel_with_fallback(f)

            if len(df) == 0:
                continue

            # 自动填充客户名到内存中的DataFrame
            if '客户(不填)' in df.columns:
                df['客户(不填)'] = df['客户(不填)'].astype(object)
                empty_mask = df['客户(不填)'].isna() | (df['客户(不填)'] == '')
                if empty_mask.any() and customer:
                    df.loc[empty_mask, '客户(不填)'] = customer
            elif customer:
                df['客户(不填)'] = customer

            df['来源文件'] = basename

            # 过滤没有图片的行（主图列既没有WPS图片也没有标准图片）
            if '__drawing_row__' in df.columns:
                pic_col = None
                for c in df.columns:
                    if str(c) == '主图':
                        pic_col = c
                        break
                if pic_col is not None:
                    valid_mask = []
                    for _, row in df.iterrows():
                        has_img = False
                        cell_val = row.get(pic_col)
                        # WPS图片检查
                        if isinstance(cell_val, str):
                            img_id = extract_img_id(cell_val)
                            if img_id and img_id in images:
                                has_img = True
                        # 标准图片检查
                        if not has_img:
                            dr = row.get('__drawing_row__')
                            if dr is not None:
                                dr_int = int(dr)
                                if pic_col in std_images_mapped and dr_int in std_images_mapped[pic_col]:
                                    has_img = True
                        valid_mask.append(has_img)
                    if not all(valid_mask):
                        removed = len(df) - sum(valid_mask)
                        df = df[valid_mask].reset_index(drop=True)
                        print(f"  [{basename}] 过滤了 {removed} 行无图片数据")

            all_data.append(df)
            print(f"  [{basename}] 读取了 {len(df)} 行")
        except Exception as e:
            print(f"  读取 {basename} 失败: {e}")

    if not all_data:
        print("没有成功读取任何数据")
        return

    combined = pd.concat(all_data, ignore_index=True)
    print(f"\n合并后总行数: {len(combined)}")

    if '商品ID' in combined.columns:
        combined['商品ID'] = combined['商品ID'].apply(
            lambda x: str(int(x)) if pd.notna(x) and str(x).replace('.','',1).isdigit() else str(x) if pd.notna(x) else ''
        )

    if '备注' not in combined.columns:
        combined['备注'] = ''

    # 识别店铺名列 - 优先找"店铺名称"，其次才考虑"客户(不填)"
    shop_col = None
    for col in combined.columns:
        col_str = str(col)
        if '店铺名称' in col_str:
            if combined[col].notna().sum() > 0:
                shop_col = col
                break

    if shop_col is None:
        for col in combined.columns:
            col_str = str(col)
            if ('客户' in col_str and '店铺' in col_str) or col_str == '客户(不填)':
                if combined[col].notna().sum() > 0:
                    shop_col = col
                    break

    if shop_col is None:
        best_col, best_count = None, 0
        for col in combined.columns:
            if str(col).startswith('Unnamed'):
                continue
            count = combined[col].apply(lambda x: isinstance(x, str) and len(x) > 3).sum()
            if count > best_count:
                best_count, best_col = count, col
        shop_col = best_col

    print(f"店铺列识别为: {shop_col}")

    # 识别单量列
    qty_col = None
    for col in combined.columns:
        col_str = str(col)
        if '补单数量' in col_str or '单量' in col_str or '数量' in col_str:
            qty_col = col
            break
    if qty_col is None:
        print("未找到单量列")
        return
    print(f"单量列识别为: {qty_col}")

    # 过滤有效数据
    valid_mask = combined[shop_col].notna() & combined[qty_col].notna()
    try:
        qty_numeric = pd.to_numeric(combined[qty_col], errors='coerce')
        valid_mask = valid_mask & (qty_numeric > 0)
    except:
        pass
    data = combined[valid_mask].copy()
    data[qty_col] = pd.to_numeric(data[qty_col], errors='coerce')
    data = data[data[qty_col] > 0].copy()

    data[shop_col] = data[shop_col].astype(str).str.replace(r'[\n\r]+', '', regex=True).str.strip()

    print(f"有效数据行数: {len(data)}")
    if len(data) == 0:
        return

    # 拆分补单数量大于1的行
    expanded_rows = []
    for _, row in data.iterrows():
        qty = int(row[qty_col])
        if qty > 1:
            for _ in range(qty):
                new_row = row.copy()
                new_row[qty_col] = 1
                expanded_rows.append(new_row)
        else:
            expanded_rows.append(row)
    data = pd.DataFrame(expanded_rows).reset_index(drop=True)
    data[qty_col] = pd.to_numeric(data[qty_col], errors='coerce')
    print(f"拆分后数据行数: {len(data)}")

    # 链接交叉排序：同一店铺内不同链接按轮询交叉排列
    data = interleave_shop_links(data, shop_col)
    print(f"链接交叉排序后数据行数: {len(data)}")

    # 解析时间
    start_h, start_m = map(int, start_time_str.split(':'))
    end_h, end_m = map(int, end_time_str.split(':'))
    base_time = datetime(2026, 1, 1, start_h, start_m)
    total_work_minutes = (end_h * 60 + end_m) - (start_h * 60 + start_m)
    if total_work_minutes <= 0:
        print("错误：结束时间必须晚于开始时间")
        return

    # 提取每个店铺的订单数，进行自动调度
    shop_orders = data.groupby(shop_col)[qty_col].sum().astype(int).to_dict()
    if num_workers < 1:
        num_workers = 1

    # 显示店铺→刷手分配方案
    shop_assignments = assign_shops_to_brushers(shop_orders, num_workers)
    print(f"\n工作时段: {start_time_str} ~ {end_time_str}，共 {total_work_minutes} 分钟，刷手数: {num_workers}")
    print("店铺分配方案（一个店铺只分配给一个刷手）：")
    for w_idx, shops in shop_assignments.items():
        total = sum(shops.values())
        shop_list = ', '.join(f"{s}({c}单)" for s, c in shops.items())
        print(f"  刷手{w_idx + 1} (共{total}单): {shop_list}")

    schedule_list = auto_schedule(start_time_str, end_time_str, shop_orders, num_workers, min_interval)

    # 按店铺整理调度结果
    shop_schedule_map = {}
    for item in schedule_list:
        shop = item['店铺']
        shop_schedule_map.setdefault(shop, []).append(item)

    # 店铺统计
    shop_stats = data.groupby(shop_col)[qty_col].sum().reset_index()
    shop_stats.columns = [shop_col, '店铺总单量']

    shop_last_time = {}
    shop_avg_interval = {}
    for shop, items in shop_schedule_map.items():
        times = sorted([it['时间分钟'] for it in items])
        shop_last_time[shop] = times[-1] if times else 0
        if len(times) > 1:
            shop_avg_interval[shop] = (times[-1] - times[0]) / (len(times) - 1)
        else:
            shop_avg_interval[shop] = 0

    shop_stats['平均间隔(分钟)'] = shop_stats[shop_col].map(shop_avg_interval)
    shop_stats['预计完成时间'] = shop_stats[shop_col].map(
        lambda s: (base_time + timedelta(minutes=shop_last_time.get(s, 0))).strftime('%H:%M')
    )

    print("\n店铺统计:")
    for _, row in shop_stats.iterrows():
        print(f"  {row[shop_col]}: 总单量={row['店铺总单量']}, 平均间隔={row['平均间隔(分钟)']:.1f}分钟, 预计完成={row['预计完成时间']}")

    # 将调度结果映射回原始数据行
    result_rows = []
    for shop_name, group in data.groupby(shop_col, sort=False):
        group = group.reset_index(drop=True)
        assignments_local = shop_schedule_map.get(shop_name, [])
        total_qty = len(group)

        for idx, (_, row) in enumerate(group.iterrows()):
            if idx < len(assignments_local):
                item = assignments_local[idx]
            else:
                item = {'时间分钟': 0, '刷手': '刷手1'}

            exec_time = base_time + timedelta(minutes=item['时间分钟'])
            new_row = row.to_dict()
            new_row['执行时间'] = exec_time.strftime('%H:%M')
            new_row['刷手'] = item['刷手']
            new_row['距离首单(分钟)'] = round(item['时间分钟'], 1)
            new_row['店铺总单量'] = total_qty
            result_rows.append(new_row)

    result_df = pd.DataFrame(result_rows)
    result_df = result_df.sort_values(by='执行时间').reset_index(drop=True)

    # 来源文件拆分核对统计
    if '来源文件' in data.columns and '来源文件' in result_df.columns:
        print("\n来源文件拆分核对:")
        before = data.groupby('来源文件').agg({qty_col: 'sum', '来源文件': 'size'})
        before.columns = ['拆分前单量', '有效行数']
        before = before.reset_index()
        after = result_df.groupby('来源文件').size().reset_index(name='拆分后行数')
        check = pd.merge(before, after, on='来源文件', how='outer').fillna(0)
        total_before = 0
        total_after = 0
        for _, row in check.iterrows():
            diff = int(row['拆分后行数'] - row['拆分前单量'])
            status = "OK" if diff == 0 else f"差异{diff}"
            total_before += row['拆分前单量']
            total_after += row['拆分后行数']
            print(f"  {row['来源文件']}: 有效行数={int(row['有效行数'])}, 拆分前单量={int(row['拆分前单量'])}, 拆分后行数={int(row['拆分后行数'])} [{status}]")
        print(f"  合计: 拆分前单量={int(total_before)}, 拆分后行数={int(total_after)} {'[OK]' if total_before == total_after else '[差异'+str(int(total_after-total_before))+']'}")

    if '序号' in result_df.columns:
        result_df.drop(columns=['序号'], inplace=True)

    if '备注' not in result_df.columns:
        result_df['备注'] = ''

    cols = list(result_df.columns)
    ordered = ['执行时间', '刷手', '备注']
    unnamed_cols = [c for c in cols if str(c).startswith('Unnamed')]
    remaining = [c for c in cols if c not in ordered and c not in unnamed_cols]
    result_df = result_df[ordered + unnamed_cols + remaining]

    result_df.insert(0, '序号', range(1, len(result_df) + 1))

    # === 统计（按文件 / 按客户 / 按店铺）===
    file_stats = []
    customer_stats = {}
    shop_stats_summary = {}

    for basename in sorted(customer_from_file.keys()):
        customer = customer_from_file[basename]
        file_data = data[data['来源文件'] == basename]
        qty_sum = file_data[qty_col].sum()
        file_stats.append({
            '文件名': basename,
            '客户名': customer,
            '补单数量': int(qty_sum) if qty_sum == int(qty_sum) else qty_sum
        })
        customer_stats[customer] = customer_stats.get(customer, 0) + qty_sum
        for shop_name, group in file_data.groupby(shop_col):
            shop_stats_summary[shop_name] = shop_stats_summary.get(shop_name, 0) + group[qty_col].sum()

    df_file_stats = pd.DataFrame(file_stats)
    df_customer_stats = pd.DataFrame(
        [{'客户名': k, '补单数量': int(v) if v == int(v) else v}
         for k, v in sorted(customer_stats.items(), key=lambda x: x[1], reverse=True)]
    )
    df_shop_stats_summary = pd.DataFrame(
        [{'店铺名': k, '补单数量': int(v) if v == int(v) else v}
         for k, v in sorted(shop_stats_summary.items(), key=lambda x: x[1], reverse=True)]
    )

    print("\n统计结果:")
    print(f"  文件数: {len(file_stats)}")
    print(f"  客户数: {len(customer_stats)}")
    print(f"  店铺数: {len(shop_stats_summary)}")
    print(f"  总单量: {int(data[qty_col].sum())}")

    # 输出路径
    if output_path is None:
        desktop = os.path.join(os.path.expanduser('~'), 'Desktop')
        output_path = os.path.join(desktop, f"汇总表_{datetime.now().strftime('%m%d_%H%M')}.xlsx")

    # 先用pandas写入Excel（排除内部使用的 __drawing_row__ 列）
    output_df = result_df.drop(columns=['__drawing_row__'], errors='ignore')
    with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
        output_df.to_excel(writer, sheet_name='汇总表', index=False)
        shop_stats.to_excel(writer, sheet_name='店铺统计', index=False)
        df_file_stats.to_excel(writer, sheet_name='按文件统计', index=False)
        df_customer_stats.to_excel(writer, sheet_name='按客户统计', index=False)
        df_shop_stats_summary.to_excel(writer, sheet_name='按店铺统计', index=False)

    # 用openpyxl插入图片并美化表格
    print("\n正在插入图片并美化表格...")
    wb = load_workbook(output_path)
    ws = wb['汇总表']

    col_idx_map = {cell.value: cell.column for cell in ws[1]}

    header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    header_font = Font(name='微软雅黑', size=11, bold=True, color='FFFFFF')
    header_align = Alignment(horizontal='center', vertical='center', wrap_text=True)

    even_fill = PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid')
    odd_fill = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')

    thin_border = Border(
        left=Side(style='thin', color='D9D9D9'),
        right=Side(style='thin', color='D9D9D9'),
        top=Side(style='thin', color='D9D9D9'),
        bottom=Side(style='thin', color='D9D9D9')
    )

    center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    left_align = Alignment(horizontal='left', vertical='center', wrap_text=True)
    right_align = Alignment(horizontal='right', vertical='center')

    pic_cols = ['主图', '后台分享码（二维码）']
    for pic_col in pic_cols:
        if pic_col in col_idx_map:
            col_letter = get_column_letter(col_idx_map[pic_col])
            ws.column_dimensions[col_letter].width = 14
    ws.row_dimensions[1].height = 30

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_align
        cell.border = thin_border

    col_widths = {
        '序号': 6,
        '执行时间': 12,
        '刷手': 10,
        '距离首单(分钟)': 14,
        '店铺总单量': 12,
        '商品标题': 50,
        '店铺名称': 35,
        '客户名(店铺)': 20,
        '客户(不填)': 15,
        '下单账号': 15,
        '支付账号': 15,
        '下单价格': 12,
        '补单数量': 12,
        '佣金': 10,
        '总计': 10,
        '来源文件': 25,
        '商品ID': 18,
        '订单号(不填)': 15,
        '备注': 30,
    }
    for col_name, width in col_widths.items():
        if col_name in col_idx_map:
            col_letter = get_column_letter(col_idx_map[col_name])
            ws.column_dimensions[col_letter].width = width

    inserted = 0
    for excel_row_idx, (df_idx, row_series) in enumerate(result_df.iterrows(), start=2):
        row_fill = even_fill if (excel_row_idx % 2 == 0) else odd_fill
        ws.row_dimensions[excel_row_idx].height = 65

        for col_idx in range(1, ws.max_column + 1):
            cell = ws.cell(row=excel_row_idx, column=col_idx)
            cell.border = thin_border
            cell.fill = row_fill
            cell.alignment = left_align

        source_file = row_series.get('来源文件')
        if not source_file:
            continue

        orig_path = None
        for fp in file_images:
            if os.path.basename(fp) == source_file:
                orig_path = fp
                break
        if not orig_path:
            continue

        images = file_images.get(orig_path, {})
        std_images = file_std_images.get(orig_path, {})
        drawing_row = row_series.get('__drawing_row__')

        for col_name in result_df.columns:
            if col_name in ('序号', '执行时间', '距离首单(分钟)', '店铺总单量', '每单间隔(分钟)', '来源文件'):
                continue
            cell_value = row_series[col_name]
            img_id = extract_img_id(cell_value)
            col_idx = col_idx_map.get(col_name)
            if not col_idx:
                continue

            cell = ws.cell(row=excel_row_idx, column=col_idx)
            has_img = False

            if img_id and img_id in images:
                cell.value = None
                try:
                    img = XLImage(io.BytesIO(images[img_id]))
                    img.width = 70
                    img.height = 70
                    cell_addr = f"{get_column_letter(col_idx)}{excel_row_idx}"
                    ws.add_image(img, cell_addr)
                    inserted += 1
                    has_img = True
                except Exception as e:
                    pass
            elif isinstance(cell_value, str) and 'DISPIMG' in cell_value:
                cell.value = None
                has_img = True

            # 标准Excel图片（非WPS格式）
            if not has_img and col_name in std_images and drawing_row is not None:
                dr = int(drawing_row)
                if dr in std_images[col_name]:
                    img_data = std_images[col_name][dr]
                    cell.value = None
                    try:
                        img = XLImage(io.BytesIO(img_data))
                        img.width = 70
                        img.height = 70
                        cell_addr = f"{get_column_letter(col_idx)}{excel_row_idx}"
                        ws.add_image(img, cell_addr)
                        inserted += 1
                    except Exception as e:
                        pass

    num_cols = ['序号', '距离首单(分钟)', '店铺总单量', '下单价格', '补单数量', '佣金', '总计']
    for col_name in num_cols:
        if col_name in col_idx_map:
            col_idx = col_idx_map[col_name]
            for row in range(1, ws.max_row + 1):
                ws.cell(row=row, column=col_idx).alignment = center_align if row == 1 else right_align

    time_cols = ['执行时间']
    for col_name in time_cols:
        if col_name in col_idx_map:
            col_idx = col_idx_map[col_name]
            for row in range(1, ws.max_row + 1):
                ws.cell(row=row, column=col_idx).alignment = center_align

    if '商品ID' in col_idx_map:
        col_idx = col_idx_map['商品ID']
        for row in range(2, ws.max_row + 1):
            cell = ws.cell(row=row, column=col_idx)
            if cell.value is not None:
                cell.number_format = '@'

    ws.freeze_panes = 'A2'

    wb.save(output_path)
    print(f"插入了 {inserted} 张图片")
    print(f"\n汇总完成！已保存到: {output_path}")
    print(f"总单数: {len(result_df)}")
    return output_path


if __name__ == '__main__':
    import sys
    default_folder = r'C:\Users\Administrator\Desktop\1'
    folder = sys.argv[1] if len(sys.argv) > 1 else default_folder
    if not os.path.exists(folder):
        print(f"文件夹不存在: {folder}")
        input("按回车键退出...")
        sys.exit(1)
    start_input = input("请输入第一单开始时间（HH:MM，默认08:00）：").strip()
    if not start_input:
        start_input = "08:00"
    end_input = input("请输入工作结束时间（HH:MM，默认23:00）：").strip()
    if not end_input:
        end_input = "23:00"
    workers_input = input("请输入刷手数量（默认1）：").strip()
    num_workers = int(workers_input) if workers_input.isdigit() else 1
    min_input = input("请输入同店铺最小间隔分钟数（默认20）：").strip()
    min_interval = float(min_input) if min_input else 20
    process_folder(folder, start_time_str=start_input, end_time_str=end_input,
                   num_workers=num_workers, min_interval=min_interval)
    input("\n按回车键退出...")

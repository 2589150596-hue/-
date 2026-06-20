import os
import re
import tkinter as tk
from tkinter import ttk, filedialog, messagebox
import pandas as pd
from openpyxl import load_workbook
from collections import defaultdict


class BudanStatsApp:
    def __init__(self, root):
        self.root = root
        self.root.title("补单统计工具")
        self.root.geometry("900x650")
        self.root.minsize(800, 500)

        self.input_dir = tk.StringVar(value=r"C:\Users\Administrator\Desktop\2")
        self.last_result = None

        self._build_ui()

    def _build_ui(self):
        # 顶部控制区
        top_frame = tk.Frame(self.root, padx=10, pady=10)
        top_frame.pack(fill=tk.X)

        tk.Label(top_frame, text="文件夹:").pack(side=tk.LEFT)
        tk.Entry(top_frame, textvariable=self.input_dir, width=60).pack(side=tk.LEFT, padx=5)
        tk.Button(top_frame, text="浏览...", command=self._browse_folder).pack(side=tk.LEFT, padx=2)
        tk.Button(top_frame, text="开始统计", bg="#4CAF50", fg="white", command=self._run_stats).pack(side=tk.LEFT, padx=10)

        # Tab页
        self.notebook = ttk.Notebook(self.root)
        self.notebook.pack(fill=tk.BOTH, expand=True, padx=10, pady=5)

        self.tab_file = self._create_tab("按文件统计")
        self.tab_customer = self._create_tab("按客户统计")
        self.tab_shop = self._create_tab("按店铺统计")

        # 底部
        bottom_frame = tk.Frame(self.root, padx=10, pady=10)
        bottom_frame.pack(fill=tk.X)

        self.status_label = tk.Label(bottom_frame, text="就绪", anchor=tk.W)
        self.status_label.pack(side=tk.LEFT)

        tk.Button(bottom_frame, text="导出Excel", command=self._export_excel).pack(side=tk.RIGHT, padx=5)

    def _create_tab(self, title):
        frame = tk.Frame(self.notebook)
        self.notebook.add(frame, text=title)

        cols = ("c1", "c2", "c3") if title == "按文件统计" else ("c1", "c2")
        tree = ttk.Treeview(frame, columns=cols, show="headings", selectmode="extended")

        if title == "按文件统计":
            tree.heading("c1", text="文件名")
            tree.heading("c2", text="客户名")
            tree.heading("c3", text="补单数量")
            tree.column("c1", width=450)
            tree.column("c2", width=100)
            tree.column("c3", width=100, anchor="center")
        elif title == "按客户统计":
            tree.heading("c1", text="客户名")
            tree.heading("c2", text="补单数量")
            tree.column("c1", width=200)
            tree.column("c2", width=150, anchor="center")
        else:
            tree.heading("c1", text="店铺名")
            tree.heading("c2", text="补单数量")
            tree.column("c1", width=350)
            tree.column("c2", width=150, anchor="center")

        vsb = ttk.Scrollbar(frame, orient="vertical", command=tree.yview)
        hsb = ttk.Scrollbar(frame, orient="horizontal", command=tree.xview)
        tree.configure(yscrollcommand=vsb.set, xscrollcommand=hsb.set)

        vsb.pack(side=tk.RIGHT, fill=tk.Y)
        hsb.pack(side=tk.BOTTOM, fill=tk.X)
        tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)

        # 右键菜单
        menu = tk.Menu(tree, tearoff=0)
        menu.add_command(label="复制选中行", command=lambda t=tree: self._copy_selected_rows(t))
        menu.add_command(label="复制单元格", command=lambda t=tree: self._copy_cell(t))
        menu.add_separator()
        menu.add_command(label="复制全部", command=lambda t=tree: self._copy_all(t))

        def popup(event):
            menu.post(event.x_root, event.y_root)
        tree.bind("<Button-3>", popup)

        frame.tree = tree
        return frame

    # ================= 复制功能 =================

    def _copy_selected_rows(self, tree):
        selected = tree.selection()
        if not selected:
            return
        lines = []
        for item in selected:
            lines.append("\t".join(str(v) for v in tree.item(item, "values")))
        text = "\n".join(lines)
        self.root.clipboard_clear()
        self.root.clipboard_append(text)

    def _copy_cell(self, tree):
        selected = tree.selection()
        if not selected:
            return
        # 取第一个选中行的第一个列值
        item = selected[0]
        values = tree.item(item, "values")
        if values:
            self.root.clipboard_clear()
            self.root.clipboard_append(str(values[0]))

    def _copy_all(self, tree):
        lines = []
        for item in tree.get_children():
            lines.append("\t".join(str(v) for v in tree.item(item, "values")))
        text = "\n".join(lines)
        self.root.clipboard_clear()
        self.root.clipboard_append(text)

    def _browse_folder(self):
        path = filedialog.askdirectory(initialdir=self.input_dir.get() or os.getcwd())
        if path:
            self.input_dir.set(path)

    # ================= 统计逻辑 =================

    @staticmethod
    def parse_customer(filename):
        name_without_ext = os.path.splitext(filename)[0]
        parts = name_without_ext.split("-", 2)
        if len(parts) >= 2:
            candidate = parts[1].strip()
            m = re.match(r"(\d+\.\d+)?(早补|晚补)?(.+)", candidate)
            if m:
                possible = m.group(3).strip()
                if possible:
                    return possible
            return candidate
        return "未知客户"

    @staticmethod
    def extract_shop_from_filename(filename):
        name_no_ext = os.path.splitext(filename)[0]
        parts = name_no_ext.split("-", 2)
        if len(parts) >= 3:
            return BudanStatsApp.clean_shop_name(parts[2])
        return "未知店铺"

    @staticmethod
    def clean_shop_name(name):
        if not name:
            return "未知店铺"
        name = str(name).strip()
        name = re.sub(r"\.xlsx", "", name, flags=re.IGNORECASE)
        name = re.sub(r"\s*\(\d+\)\s*$", "", name)
        name = re.sub(r"-\d+-?\s*$", "", name)
        name = re.sub(r"\d+\.\d+.*$", "", name)
        name = re.sub(r"\d{4}[\.\s]+\d+[\.\s]+\d+.*$", "", name)
        return name.strip()

    @staticmethod
    def normalize_shop_name(name):
        if not name or str(name).strip() == "":
            return "未知店铺"
        name = str(name).strip()
        name = re.sub(r"[\(\)（）]", "", name)
        return name.strip()

    @staticmethod
    def find_column_index(headers_r1, headers_r2, keywords):
        def search(headers):
            for keyword in keywords:
                for idx, h in enumerate(headers):
                    if h and keyword in str(h).strip():
                        return idx
            return None
        idx = search(headers_r1)
        if idx is not None:
            return idx
        if headers_r2:
            idx = search(headers_r2)
            if idx is not None:
                return idx
        return None

    def read_excel_data(self, filepath):
        filename = os.path.basename(filepath)
        customer = self.parse_customer(filename)
        shop_from_filename = self.extract_shop_from_filename(filename)

        try:
            wb = load_workbook(filepath, data_only=True)
            ws = wb.active
        except Exception as e:
            return customer, 0, {}, f"无法打开文件: {e}"

        rows = list(ws.iter_rows(min_row=1, max_row=3, values_only=True))
        if not rows:
            return customer, 0, {}, "空文件"

        headers_r1 = [str(cell).strip() if cell is not None else None for cell in rows[0]]
        headers_r2 = None
        if len(rows) >= 2:
            has_header_keywords = any(
                cell is not None and isinstance(cell, str)
                and any(k in cell for k in ["数量", "价格", "时间"])
                for cell in rows[1]
            )
            if has_header_keywords:
                headers_r2 = [str(cell).strip() if cell is not None else None for cell in rows[1]]

        qty_col = self.find_column_index(headers_r1, headers_r2, ["补单数量", "补数量", "下单数量", "数量"])
        if qty_col is None:
            return customer, 0, {}, f"未找到数量列"

        shop_col = self.find_column_index(headers_r1, headers_r2, ["店铺名称", "店铺", "客户(不填)", "客户名(不填)"])
        total_col = self.find_column_index(headers_r1, headers_r2, ["总计", "合计", "总"])

        data_start_row = 3 if headers_r2 else 2

        shop_stats = defaultdict(int)
        total = 0

        for row in ws.iter_rows(min_row=data_start_row, values_only=True):
            if not row or all(cell is None for cell in row):
                continue

            # 跳过汇总行
            if total_col is not None and total_col < len(row):
                total_val = row[total_col]
                qty_val_check = row[qty_col] if qty_col < len(row) else None
                if total_val == 0 and total_val is not False:
                    try:
                        float(qty_val_check)
                        continue
                    except (ValueError, TypeError):
                        pass

            skip = False
            for cell in row:
                if cell is not None and isinstance(cell, str) and any(k in cell for k in ["合计", "总计", "汇总"]):
                    skip = True
                    break
            if skip:
                continue

            qty_val = row[qty_col] if qty_col < len(row) else None
            if qty_val is None:
                continue
            try:
                qty = float(qty_val)
            except (ValueError, TypeError):
                continue

            total += qty

            shop_name = None
            if shop_col is not None and shop_col < len(row):
                shop_name = row[shop_col]
            if not shop_name or str(shop_name).strip() == "":
                shop_name = shop_from_filename
            else:
                shop_name = self.clean_shop_name(shop_name)

            normalized = self.normalize_shop_name(shop_name)
            shop_stats[normalized] += qty

        return customer, total, dict(shop_stats), None

    # ================= 运行与展示 =================

    def _run_stats(self):
        input_dir = self.input_dir.get()
        if not os.path.isdir(input_dir):
            messagebox.showerror("错误", "请选择有效的文件夹路径")
            return

        files = [f for f in os.listdir(input_dir)
                 if f.lower().endswith(".xlsx") and not f.startswith("~$")
                 and "统计结果" not in f and "统计" not in f]

        if not files:
            messagebox.showinfo("提示", "未找到Excel文件")
            return

        self.status_label.config(text="正在统计...")
        self.root.update_idletasks()

        file_results = []
        customer_stats = defaultdict(float)
        shop_stats = defaultdict(float)
        errors = []

        for f in sorted(files, key=lambda x: int(re.findall(r"^\d+", x)[0]) if re.findall(r"^\d+", x) else 999):
            filepath = os.path.join(input_dir, f)
            customer, total, shops, err = self.read_excel_data(filepath)
            total_int = int(total) if isinstance(total, float) and total.is_integer() else total
            if err:
                file_results.append({"文件名": f, "客户名": customer, "补单数量": 0, "备注": err})
                errors.append(f"{f}: {err}")
            else:
                file_results.append({"文件名": f, "客户名": customer, "补单数量": total_int, "备注": ""})
                customer_stats[customer] += total_int
                for shop, qty in shops.items():
                    shop_stats[shop] += int(qty) if isinstance(qty, float) and qty.is_integer() else qty

        # 填充Treeview
        self._fill_tree(self.tab_file.tree, file_results, ["文件名", "客户名", "补单数量"])
        self._fill_tree(self.tab_customer.tree,
                        [{"客户名": k, "补单数量": v} for k, v in sorted(customer_stats.items(), key=lambda x: x[1], reverse=True)],
                        ["客户名", "补单数量"])
        self._fill_tree(self.tab_shop.tree,
                        [{"店铺名": k, "补单数量": v} for k, v in sorted(shop_stats.items(), key=lambda x: x[1], reverse=True)],
                        ["店铺名", "补单数量"])

        self.last_result = {
            "file_results": file_results,
            "customer_stats": dict(customer_stats),
            "shop_stats": dict(shop_stats),
            "input_dir": input_dir
        }

        msg = f"统计完成: {len(files)}个文件"
        if errors:
            msg += f", {len(errors)}个文件出错"
        self.status_label.config(text=msg)

        if errors:
            messagebox.showwarning("部分文件出错", "\n".join(errors[:5]) + ("\n..." if len(errors) > 5 else ""))

    def _fill_tree(self, tree, data, keys):
        tree.delete(*tree.get_children())
        for row in data:
            vals = []
            for k in keys:
                v = row.get(k, "")
                if k == "补单数量" and isinstance(v, float):
                    v = int(v)
                vals.append(v)
            tree.insert("", tk.END, values=vals)

    def _export_excel(self):
        if not self.last_result:
            messagebox.showwarning("提示", "请先执行统计")
            return

        path = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            initialfile="统计结果.xlsx",
            filetypes=[("Excel文件", "*.xlsx")]
        )
        if not path:
            return

        df_file = pd.DataFrame(self.last_result["file_results"])
        df_customer = pd.DataFrame(
            [{"客户名": k, "补单数量": v} for k, v in sorted(self.last_result["customer_stats"].items(), key=lambda x: x[1], reverse=True)]
        )
        df_shop = pd.DataFrame(
            [{"店铺名": k, "补单数量": v} for k, v in sorted(self.last_result["shop_stats"].items(), key=lambda x: x[1], reverse=True)]
        )

        try:
            with pd.ExcelWriter(path, engine="openpyxl") as writer:
                df_file.to_excel(writer, sheet_name="按文件统计", index=False)
                df_customer.to_excel(writer, sheet_name="按客户统计", index=False)
                df_shop.to_excel(writer, sheet_name="按店铺统计", index=False)
            messagebox.showinfo("成功", f"已导出到:\n{path}")
        except Exception as e:
            messagebox.showerror("导出失败", str(e))


def main():
    root = tk.Tk()
    app = BudanStatsApp(root)
    root.mainloop()


if __name__ == "__main__":
    main()

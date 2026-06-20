import os
import re
import copy
import tkinter as tk
from tkinter import ttk, filedialog, messagebox
import pandas as pd
from openpyxl import load_workbook
from collections import defaultdict


# ===================== 核心统计逻辑 =====================

def parse_customer(filename):
    name_without_ext = os.path.splitext(filename)[0]
    parts = name_without_ext.split("-", 2)
    if len(parts) >= 2:
        candidate = parts[1].strip()
        m = re.match(r"(\d+\.\d+)?(早补|晚补)?(.+)", candidate)
        if m:
            possible = m.group(3).strip()
            if possible:
                return possible
        return candidate
    return "未知客户"


def extract_shop_from_filename(filename):
    name_no_ext = os.path.splitext(filename)[0]
    parts = name_no_ext.split("-", 2)
    if len(parts) >= 3:
        return clean_shop_name(parts[2])
    return "未知店铺"


def clean_shop_name(name):
    if not name:
        return "未知店铺"
    name = str(name).strip()
    name = re.sub(r"\.xlsx", "", name, flags=re.IGNORECASE)
    name = re.sub(r"\s*\(\d+\)\s*$", "", name)
    name = re.sub(r"-\d+-?\s*$", "", name)
    name = re.sub(r"\d+\.\d+.*$", "", name)
    name = re.sub(r"\d{4}[\.\s]+\d+[\.\s]+\d+.*$", "", name)
    return name.strip()


def normalize_shop_name(name):
    if not name or str(name).strip() == "":
        return "未知店铺"
    name = str(name).strip()
    name = re.sub(r"[\(\)（）]", "", name)
    return name.strip()


def find_column_index(headers_r1, headers_r2, keywords):
    def search(headers):
        for keyword in keywords:
            for idx, h in enumerate(headers):
                if h and keyword in str(h).strip():
                    return idx
        return None
    idx = search(headers_r1)
    if idx is not None:
        return idx
    if headers_r2:
        idx = search(headers_r2)
        if idx is not None:
            return idx
    return None


def read_excel_data(filepath):
    """
    读取单个Excel，返回字典包含：
    customer, total, shop_stats, error, data_hash, data_rows, file_content_hash
    """
    filename = os.path.basename(filepath)
    customer = parse_customer(filename)
    shop_from_filename = extract_shop_from_filename(filename)

    result = {
        "filename": filename,
        "customer": customer,
        "total": 0,
        "shop_stats": {},
        "error": None,
        "data_rows": 0,
        "content_hash": None,
        "filepath": filepath,
    }

    try:
        wb = load_workbook(filepath, data_only=True)
        ws = wb.active
    except Exception as e:
        result["error"] = f"无法打开文件: {e}"
        return result

    rows = list(ws.iter_rows(min_row=1, max_row=3, values_only=True))
    if not rows:
        result["error"] = "空文件"
        return result

    headers_r1 = [str(cell).strip() if cell is not None else None for cell in rows[0]]
    headers_r2 = None
    if len(rows) >= 2:
        has_header_keywords = any(
            cell is not None and isinstance(cell, str)
            and any(k in cell for k in ["数量", "价格", "时间"])
            for cell in rows[1]
        )
        if has_header_keywords:
            headers_r2 = [str(cell).strip() if cell is not None else None for cell in rows[1]]

    qty_col = find_column_index(headers_r1, headers_r2, ["补单数量", "补数量", "下单数量", "数量"])
    if qty_col is None:
        result["error"] = "未找到数量列"
        return result

    shop_col = find_column_index(headers_r1, headers_r2, ["店铺名称", "店铺", "客户(不填)", "客户名(不填)"])
    total_col = find_column_index(headers_r1, headers_r2, ["总计", "合计", "总"])
    customer_col = find_column_index(headers_r1, headers_r2, ["客户(不填)", "客户名(不填)", "客户"])

    data_start_row = 3 if headers_r2 else 2

    shop_stats = defaultdict(int)
    total = 0
    data_rows = 0
    content_tuples = []

    for row in ws.iter_rows(min_row=data_start_row, values_only=True):
        if not row or all(cell is None for cell in row):
            continue

        if total_col is not None and total_col < len(row):
            total_val = row[total_col]
            qty_val_check = row[qty_col] if qty_col < len(row) else None
            if total_val == 0 and total_val is not False:
                try:
                    float(qty_val_check)
                    continue
                except (ValueError, TypeError):
                    pass

        skip = False
        for cell in row:
            if cell is not None and isinstance(cell, str) and any(k in cell for k in ["合计", "总计", "汇总"]):
                skip = True
                break
        if skip:
            continue

        qty_val = row[qty_col] if qty_col < len(row) else None
        if qty_val is None:
            continue
        try:
            qty = float(qty_val)
        except (ValueError, TypeError):
            continue

        total += qty
        data_rows += 1

        # 收集内容用于hash（用于异常检测：完全相同的表格）
        row_tuple = tuple(str(cell) if cell is not None else "" for cell in row)
        content_tuples.append(row_tuple)

        shop_name = None
        if shop_col is not None and shop_col < len(row):
            shop_name = row[shop_col]
        if not shop_name or str(shop_name).strip() == "":
            shop_name = shop_from_filename
        else:
            shop_name = clean_shop_name(shop_name)

        normalized = normalize_shop_name(shop_name)
        shop_stats[normalized] += qty

    result["total"] = total
    result["shop_stats"] = dict(shop_stats)
    result["data_rows"] = data_rows
    result["content_hash"] = hash(tuple(content_tuples)) if content_tuples else None
    result["customer_col"] = customer_col
    result["data_start_row"] = data_start_row
    return result


# ===================== 异常检测 =====================

def detect_duplicates(results):
    """检测内容完全相同的文件，返回 [(文件1, 文件2), ...]"""
    hash_map = {}
    duplicates = []
    for r in results:
        h = r.get("content_hash")
        if h is None:
            continue
        if h in hash_map:
            duplicates.append((hash_map[h]["filename"], r["filename"]))
        else:
            hash_map[h] = r
    return duplicates


def detect_invalid_files(results):
    """检测异常文件：有错误、0数据行、总单量0"""
    invalids = []
    for r in results:
        if r.get("error"):
            invalids.append({"filename": r["filename"], "reason": r["error"]})
        elif r.get("data_rows", 0) == 0:
            invalids.append({"filename": r["filename"], "reason": "无有效数据行"})
        elif r.get("total", 0) == 0:
            invalids.append({"filename": r["filename"], "reason": "补单数量为0"})
    return invalids


# ===================== 刷手分配算法 =====================

def assign_shops_to_brushers(shop_stats, brusher_list):
    """
    改进版LPT贪心分配
    输入: {店铺名: 单量}, [刷手1, 刷手2, ...]
    输出: {刷手名: [{店铺, 单量}, ...]}
    """
    if not shop_stats or not brusher_list:
        return {}

    n = len(brusher_list)
    # 按单量降序排序
    shops = sorted(shop_stats.items(), key=lambda x: x[1], reverse=True)

    # 初始化刷手状态
    brusher_state = {name: {"shops": [], "total": 0} for name in brusher_list}

    for shop_name, qty in shops:
        # 找出当前总单量最少的刷手
        min_total = min(state["total"] for state in brusher_state.values())
        candidates = [name for name, state in brusher_state.items() if state["total"] == min_total]

        if len(candidates) == 1:
            chosen = candidates[0]
        else:
            # 平局时：选择已有店铺单量平均值与当前店铺最接近的刷手
            best = candidates[0]
            best_diff = float("inf")
            for name in candidates:
                existing = brusher_state[name]["shops"]
                if existing:
                    avg = sum(s["qty"] for s in existing) / len(existing)
                    diff = abs(avg - qty)
                else:
                    diff = float("inf")  # 空的优先，但这里所有都是平局
                if diff < best_diff:
                    best_diff = diff
                    best = name
            chosen = best

        brusher_state[chosen]["shops"].append({"店铺": shop_name, "单量": int(qty)})
        brusher_state[chosen]["total"] += qty

    # 整理输出格式
    output = {}
    for name in brusher_list:
        output[name] = brusher_state[name]["shops"]
    return output


# ===================== 客户名写入 =====================

def apply_customer_names(results, input_dir):
    """将客户名写入原始Excel的'客户(不填)'列（仅当该列为空时）"""
    applied = []
    failed = []
    for r in results:
        if r.get("error"):
            continue
        filepath = r["filepath"]
        customer = r["customer"]
        customer_col = r.get("customer_col")
        if customer_col is None:
            failed.append(f"{r['filename']}: 未找到客户列")
            continue

        try:
            wb = load_workbook(filepath)
            ws = wb.active
            start_row = r.get("data_start_row", 2)

            filled = 0
            for row in ws.iter_rows(min_row=start_row, values_only=False):
                cell = row[customer_col]
                if cell.value is None or str(cell.value).strip() == "":
                    cell.value = customer
                    filled += 1

            if filled > 0:
                wb.save(filepath)
                applied.append(f"{r['filename']}: 填充{filled}行")
            else:
                applied.append(f"{r['filename']}: 无需填充")
        except Exception as e:
            failed.append(f"{r['filename']}: {e}")

    return applied, failed


# ===================== GUI =====================

class BudanStatsAppV2:
    def __init__(self, root):
        self.root = root
        self.root.title("补单统计工具 v2")
        self.root.geometry("1000x700")
        self.root.minsize(900, 550)

        self.input_dir = tk.StringVar(value=r"C:\Users\Administrator\Desktop\2")
        self.brushers_text = tk.StringVar()
        self.last_results = None
        self.last_assignment = None

        self._build_ui()

    def _build_ui(self):
        # 顶部控制区
        top_frame = tk.Frame(self.root, padx=10, pady=10)
        top_frame.pack(fill=tk.X)

        # 文件夹
        row1 = tk.Frame(top_frame)
        row1.pack(fill=tk.X, pady=2)
        tk.Label(row1, text="文件夹:").pack(side=tk.LEFT)
        tk.Entry(row1, textvariable=self.input_dir, width=60).pack(side=tk.LEFT, padx=5)
        tk.Button(row1, text="浏览...", command=self._browse_folder).pack(side=tk.LEFT, padx=2)

        # 刷手名单
        row2 = tk.Frame(top_frame)
        row2.pack(fill=tk.X, pady=2)
        tk.Label(row2, text="刷手名单:").pack(side=tk.LEFT)
        self.brushers_entry = tk.Text(row2, height=3, width=40, font=("微软雅黑", 10))
        self.brushers_entry.pack(side=tk.LEFT, padx=5)
        tk.Label(row2, text="（每行一个刷手名）", fg="gray").pack(side=tk.LEFT)

        # 按钮行
        row3 = tk.Frame(top_frame)
        row3.pack(fill=tk.X, pady=5)
        tk.Button(row3, text="开始统计", bg="#4CAF50", fg="white", width=12, command=self._run_stats).pack(side=tk.LEFT, padx=5)
        tk.Button(row3, text="应用客户名", bg="#2196F3", fg="white", width=12, command=self._apply_customer_names).pack(side=tk.LEFT, padx=5)
        tk.Button(row3, text="导出Excel", width=12, command=self._export_excel).pack(side=tk.LEFT, padx=5)

        # Tab页
        self.notebook = ttk.Notebook(self.root)
        self.notebook.pack(fill=tk.BOTH, expand=True, padx=10, pady=5)

        self.tab_file = self._create_tab("按文件统计", ["文件名", "客户名", "补单数量"], [450, 100, 100])
        self.tab_customer = self._create_tab("按客户统计", ["客户名", "补单数量"], [300, 150])
        self.tab_shop = self._create_tab("按店铺统计", ["店铺名", "补单数量"], [450, 150])
        self.tab_assignment = self._create_tab("分配结果", ["刷手", "店铺", "单量"], [120, 350, 100])
        self.tab_anomaly = self._create_tab("异常报告", ["文件名", "异常原因"], [450, 400])

        # 底部
        bottom_frame = tk.Frame(self.root, padx=10, pady=10)
        bottom_frame.pack(fill=tk.X)
        self.status_label = tk.Label(bottom_frame, text="就绪", anchor=tk.W)
        self.status_label.pack(side=tk.LEFT)

    def _create_tab(self, title, columns, widths):
        frame = tk.Frame(self.notebook)
        self.notebook.add(frame, text=title)

        col_ids = [f"c{i}" for i in range(len(columns))]
        tree = ttk.Treeview(frame, columns=col_ids, show="headings", selectmode="extended")

        for i, (col_name, width) in enumerate(zip(columns, widths)):
            tree.heading(f"c{i}", text=col_name)
            anchor = "center" if col_name in ["补单数量", "单量"] else "w"
            tree.column(f"c{i}", width=width, anchor=anchor)

        vsb = ttk.Scrollbar(frame, orient="vertical", command=tree.yview)
        hsb = ttk.Scrollbar(frame, orient="horizontal", command=tree.xview)
        tree.configure(yscrollcommand=vsb.set, xscrollcommand=hsb.set)
        vsb.pack(side=tk.RIGHT, fill=tk.Y)
        hsb.pack(side=tk.BOTTOM, fill=tk.X)
        tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)

        # 右键菜单
        menu = tk.Menu(tree, tearoff=0)
        menu.add_command(label="复制选中行", command=lambda t=tree: self._copy_selected_rows(t))
        menu.add_command(label="复制单元格", command=lambda t=tree: self._copy_cell(t))
        menu.add_separator()
        menu.add_command(label="复制全部", command=lambda t=tree: self._copy_all(t))
        tree.bind("<Button-3>", lambda e, m=menu: m.post(e.x_root, e.y_root))

        frame.tree = tree
        return frame

    def _browse_folder(self):
        path = filedialog.askdirectory(initialdir=self.input_dir.get() or os.getcwd())
        if path:
            self.input_dir.set(path)

    def _copy_selected_rows(self, tree):
        selected = tree.selection()
        if not selected:
            return
        lines = ["\t".join(str(v) for v in tree.item(item, "values")) for item in selected]
        self.root.clipboard_clear()
        self.root.clipboard_append("\n".join(lines))

    def _copy_cell(self, tree):
        selected = tree.selection()
        if selected:
            values = tree.item(selected[0], "values")
            if values:
                self.root.clipboard_clear()
                self.root.clipboard_append(str(values[0]))

    def _copy_all(self, tree):
        lines = ["\t".join(str(v) for v in tree.item(item, "values")) for item in tree.get_children()]
        self.root.clipboard_clear()
        self.root.clipboard_append("\n".join(lines))

    def _run_stats(self):
        input_dir = self.input_dir.get()
        if not os.path.isdir(input_dir):
            messagebox.showerror("错误", "请选择有效的文件夹路径")
            return

        files = [f for f in os.listdir(input_dir)
                 if f.lower().endswith(".xlsx") and not f.startswith("~$")
                 and "统计结果" not in f and "统计" not in f and ".backup" not in f.lower()]

        if not files:
            messagebox.showinfo("提示", "未找到Excel文件")
            return

        self.status_label.config(text="正在统计...")
        self.root.update_idletasks()

        results = []
        file_results = []
        customer_stats = defaultdict(float)
        shop_stats = defaultdict(float)

        for f in sorted(files, key=lambda x: int(re.findall(r"^\d+", x)[0]) if re.findall(r"^\d+", x) else 999):
            filepath = os.path.join(input_dir, f)
            r = read_excel_data(filepath)
            results.append(r)
            if r.get("error"):
                file_results.append({"文件名": f, "客户名": r["customer"], "补单数量": 0})
            else:
                total_int = int(r["total"]) if isinstance(r["total"], float) and r["total"].is_integer() else r["total"]
                file_results.append({"文件名": f, "客户名": r["customer"], "补单数量": total_int})
                customer_stats[r["customer"]] += total_int
                for shop, qty in r["shop_stats"].items():
                    shop_stats[shop] += int(qty) if isinstance(qty, float) and qty.is_integer() else qty

        # 异常检测
        duplicates = detect_duplicates(results)
        invalids = detect_invalid_files(results)

        # 刷手分配
        brusher_names = [line.strip() for line in self.brushers_entry.get("1.0", tk.END).strip().split("\n") if line.strip()]
        assignment = {}
        if brusher_names and shop_stats:
            assignment = assign_shops_to_brushers(shop_stats, brusher_names)

        # 填充表格
        self._fill_tree(self.tab_file.tree, file_results, ["文件名", "客户名", "补单数量"])
        self._fill_tree(self.tab_customer.tree,
                        [{"客户名": k, "补单数量": v} for k, v in sorted(customer_stats.items(), key=lambda x: x[1], reverse=True)],
                        ["客户名", "补单数量"])
        self._fill_tree(self.tab_shop.tree,
                        [{"店铺名": k, "补单数量": v} for k, v in sorted(shop_stats.items(), key=lambda x: x[1], reverse=True)],
                        ["店铺名", "补单数量"])

        # 分配结果
        assign_rows = []
        for brusher, shops in assignment.items():
            for s in shops:
                assign_rows.append({"刷手": brusher, "店铺": s["店铺"], "单量": s["单量"]})
            assign_rows.append({"刷手": brusher, "店铺": "【小计】", "单量": int(sum(s["单量"] for s in shops))})
        self._fill_tree(self.tab_assignment.tree, assign_rows, ["刷手", "店铺", "单量"])

        # 异常报告
        anomaly_rows = []
        for dup in duplicates:
            anomaly_rows.append({"文件名": f"{dup[0]} <-> {dup[1]}", "异常原因": "内容完全相同的文件"})
        for inv in invalids:
            anomaly_rows.append({"文件名": inv["filename"], "异常原因": inv["reason"]})
        self._fill_tree(self.tab_anomaly.tree, anomaly_rows, ["文件名", "异常原因"])

        # 保存结果
        self.last_results = results
        self.last_assignment = assignment
        self.last_file_results = file_results
        self.last_customer_stats = dict(customer_stats)
        self.last_shop_stats = dict(shop_stats)
        self.last_anomalies = anomaly_rows

        msg = f"统计完成: {len(files)}个文件"
        if anomaly_rows:
            msg += f", 发现{len(anomaly_rows)}个异常"
        self.status_label.config(text=msg)

        if anomaly_rows:
            self.notebook.select(self.tab_anomaly)
            if duplicates:
                dup_msg = "\n".join([f"{a} 和 {b}" for a, b in duplicates])
                messagebox.showwarning("发现重复文件", f"以下文件内容完全相同:\n{dup_msg}")

    def _fill_tree(self, tree, data, keys):
        tree.delete(*tree.get_children())
        for row in data:
            vals = []
            for k in keys:
                v = row.get(k, "")
                if k in ["补单数量", "单量"] and isinstance(v, float):
                    v = int(v)
                vals.append(v)
            tree.insert("", tk.END, values=vals)

    def _apply_customer_names(self):
        if not self.last_results:
            messagebox.showwarning("提示", "请先执行统计")
            return

        if not messagebox.askyesno("确认", "将把从文件名提取的客户名写入原始Excel的'客户(不填)'列。\n是否继续？"):
            return

        applied, failed = apply_customer_names(self.last_results, self.input_dir.get())
        msg = ""
        if applied:
            msg += "成功:\n" + "\n".join(applied[:10])
            if len(applied) > 10:
                msg += f"\n...等共{len(applied)}个文件"
        if failed:
            msg += "\n\n失败:\n" + "\n".join(failed[:10])
            if len(failed) > 10:
                msg += f"\n...等共{len(failed)}个文件"
        messagebox.showinfo("应用结果", msg or "无需处理")

    def _export_excel(self):
        if not hasattr(self, "last_file_results"):
            messagebox.showwarning("提示", "请先执行统计")
            return

        path = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            initialfile="统计结果.xlsx",
            filetypes=[("Excel文件", "*.xlsx")]
        )
        if not path:
            return

        try:
            df_file = pd.DataFrame(self.last_file_results)
            df_customer = pd.DataFrame(
                [{"客户名": k, "补单数量": v} for k, v in sorted(self.last_customer_stats.items(), key=lambda x: x[1], reverse=True)]
            )
            df_shop = pd.DataFrame(
                [{"店铺名": k, "补单数量": v} for k, v in sorted(self.last_shop_stats.items(), key=lambda x: x[1], reverse=True)]
            )
            df_anomaly = pd.DataFrame(self.last_anomalies)

            # 分配结果表格
            assign_rows = []
            for brusher, shops in self.last_assignment.items():
                for s in shops:
                    assign_rows.append({"刷手": brusher, "店铺": s["店铺"], "单量": s["单量"]})
            df_assign = pd.DataFrame(assign_rows)

            with pd.ExcelWriter(path, engine="openpyxl") as writer:
                df_file.to_excel(writer, sheet_name="按文件统计", index=False)
                df_customer.to_excel(writer, sheet_name="按客户统计", index=False)
                df_shop.to_excel(writer, sheet_name="按店铺统计", index=False)
                if not df_assign.empty:
                    df_assign.to_excel(writer, sheet_name="分配结果", index=False)
                if not df_anomaly.empty:
                    df_anomaly.to_excel(writer, sheet_name="异常报告", index=False)

            messagebox.showinfo("成功", f"已导出到:\n{path}")
        except Exception as e:
            messagebox.showerror("导出失败", str(e))


def main():
    root = tk.Tk()
    app = BudanStatsAppV2(root)
    root.mainloop()


if __name__ == "__main__":
    main()
